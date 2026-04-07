#!/usr/bin/env python3
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
import os
import sys
import glob
import json
from collections import defaultdict
import webbrowser

if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

EXCEL_FILE = os.path.join(BASE_DIR, "sts2_cards.xlsx")

CLASS_COLORS = {
    "IRONCLAD": "#FF8B8B",
    "SILENT": "#8BFF8B",
    "DEFECT": "#8B8BFF",
    "NECROBINDER": "#FF8BFF",
    "REGENT": "#FFFF8B",
    "COLORLESS": "#E0E0E0",
}

CARD_CLASSIFICATIONS = {
    "ABRASIVE": "SILENT",
    "ACCELERANT": "SILENT",
    "ACCURACY": "SILENT",
    "ACROBATICS": "SILENT",
    "ADAPTIVE_STRIKE": "DEFECT",
    "ADRENALINE": "SILENT",
    "AFTERIMAGE": "SILENT",
    "AGGRESSION": "IRONCLAD",
    "ALCHEMIZE": "COLORLESS",
    "ALL_FOR_ONE": "DEFECT",
    "ANGER": "IRONCLAD",
    "ANOINTED": "DEFECT",
    "ANTICIPATE": "DEFECT",
    "APOTHEOSIS": "COLORLESS",
    "ARBITER": "NECROBINDER",
    "ARMAMENTS": "IRONCLAD",
    "ARTIFACT": "DEFECT",
    "ASH": "REGENT",
    "BACKFLIP": "SILENT",
    "BALANCE": "REGENT",
    "BANDAGE_UP": "COLORLESS",
    "BANE": "SILENT",
    "BASH": "IRONCLAD",
    "BEAM_CELL": "DEFECT",
    "BEAT_DOWN": "IRONCLAD",
    "BECKON": "NECROBINDER",
    "BERSERK": "IRONCLAD",
    "BESTIAL": "REGENT",
    "BET": "COLORLESS",
    "BEWARE": "REGENT",
    "BLASPHEMY": "COLORLESS",
    "BLIND": "SILENT",
    "BLINK": "DEFECT",
    "BLOCK": "IRONCLAD",
    "BLOODLETTING": "IRONCLAD",
    "BLOOD_PUNCH": "IRONCLAD",
    "BODY_SLAM": "IRONCLAD",
    "BOLT": "DEFECT",
    "BONE": "NECROBINDER",
    "BOOST": "DEFECT",
    "BOTTLE": "REGENT",
    "BRAMBLE": "NECROBINDER",
    "BREACH": "DEFECT",
    "BREAKTHROUGH": "DEFECT",
    "BRUTALITY": "IRONCLAD",
    "BURNING_PACT": "IRONCLAD",
    "CALCIFY": "NECROBINDER",
    "CALTROPS": "COLORLESS",
    "CASCADE": "IRONCLAD",
    "CATALYST": "SILENT",
    "CHAIN": "DEFECT",
    "CHARGED_BOLT": "DEFECT",
    "CHARGE": "REGENT",
    "CHARGE_BATTERY": "DEFECT",
    "CHEAT": "NECROBINDER",
    "CHOP": "SILENT",
    "CLAW": "DEFECT",
    "CLEAR_THE_WAY": "REGENT",
    "CLOTHESLINE": "IRONCLAD",
    "CLUMSY": "COLORLESS",
    "COALESCE": "NECROBINDER",
    "COLD_SNAP": "DEFECT",
    "COMBUST": "IRONCLAD",
    "CONCENTRATE": "SILENT",
    "CONDENSER": "DEFECT",
    "CONDUIT": "DEFECT",
    "CONSECRATE": "IRONCLAD",
    "CONTEMPLATE": "DEFECT",
    "COOLHEAD": "DEFECT",
    "CORE_SURGE": "DEFECT",
    "CORRUPT": "NECROBINDER",
    "COUNTER": "IRONCLAD",
    "CRACKLE": "DEFECT",
    "CRASH": "DEFECT",
    "CRESCENDO": "SILENT",
    "CRUSH": "IRONCLAD",
    "CURSE_OF_THE_BELL": "COLORLESS",
    "CUT": "SILENT",
    "DAGGER_THROW": "SILENT",
    "DARK_EMBRACE": "NECROBINDER",
    "DARK_SHACKLES": "IRONCLAD",
    "DAZED": "NECROBINDER",
    "DEATH_HARROW": "NECROBINDER",
    "DEBILITATE": "NECROBINDER",
    "DEFILE": "NECROBINDER",
    "DEFEND": "IRONCLAD",
    "DEFEND_REGENT": "REGENT",
    "DEFEND_NECROBINDER": "NECROBINDER",
    "DEFLECT": "SILENT",
    "DEMON_FORM": "IRONCLAD",
    "DEVASTATE": "REGENT",
    "DIRE_DOSIS": "DEFECT",
    "DISARM": "IRONCLAD",
    "DISCOVERY": "COLORLESS",
    "DISTRIBUTE": "DEFECT",
    "DODGE_AND_WEAVE": "SILENT",
    "DRAFT": "REGENT",
    "DRAIN": "NECROBINDER",
    "DRAMATIC_ENTRANCE": "REGENT",
    "DREAM_COMPOSER": "DEFECT",
    "DUALCAST": "DEFECT",
    "DYNAMO": "DEFECT",
    "EATER": "DEFECT",
    "EGG": "REGENT",
    "ELECTRODYNAMICS": "DEFECT",
    "EMANATE": "REGENT",
    "EMBED": "DEFECT",
    "EMPTY_BODY": "SILENT",
    "EMPTY_FIST": "IRONCLAD",
    "ENDURE": "IRONCLAD",
    "ENERGIZE": "DEFECT",
    "ENLIGHTEN": "DEFECT",
    "ENRAGE": "IRONCLAD",
    "ENVOY": "REGENT",
    "EQUILIBRIUM": "DEFECT",
    "EVISCERATE": "SILENT",
    "EVOKE_ORB": "DEFECT",
    "EXE": "DEFECT",
    "EXPERTISE": "DEFECT",
    "EXPLODE": "DEFECT",
    "EXPOSURE": "DEFECT",
    "EXTORT": "NECROBINDER",
    "FEAR": "NECROBINDER",
    "FEED": "IRONCLAD",
    "FEEL_NO_PAIN": "IRONCLAD",
    "FIDGET": "SILENT",
    "FINESSE": "COLORLESS",
    "FINISH": "SILENT",
    "FLAME_BURST": "DEFECT",
    "FLAMETHROWER": "DEFECT",
    "FLAP": "REGENT",
    "FLASK": "COLORLESS",
    "FLECHETTES": "SILENT",
    "FLOAT": "DEFECT",
    "FOCUS": "DEFECT",
    "FOLLY": "COLORLESS",
    "FORGE": "IRONCLAD",
    "FORM": "DEFECT",
    "FORTIFY": "NECROBINDER",
    "FURY": "IRONCLAD",
    "GAIN_STRENGTH": "DEFECT",
    "GASH": "IRONCLAD",
    "GHOSTLY": "SILENT",
    "GLACIER": "DEFECT",
    "GLASS": "DEFECT",
    "GOBLIN": "REGENT",
    "GOLDEN_IDOL": "COLORLESS",
    "GREED": "COLORLESS",
    "GRIP": "DEFECT",
    "GUILLOTINE": "IRONCLAD",
    "GUST": "DEFECT",
    "HAUNT": "NECROBINDER",
    "HEATSINK": "DEFECT",
    "HEAVY_BLADE": "IRONCLAD",
    "HEGEMONY": "REGENT",
    "HELIX": "DEFECT",
    "HEMOKINESIS": "IRONCLAD",
    "HEX": "DEFECT",
    "HIDDEN_Cache": "COLORLESS",
    "HIGHLIGHT": "DEFECT",
    "HODGIES_CLAW": "DEFECT",
    "HOMICIDE": "NECROBINDER",
    "IMpervious": "IRONCLAD",
    "IMPROVISED_EXPLOSION": "REGENT",
    "INFERNAL_BLAST": "IRONCLAD",
    "INFLAME": "IRONCLAD",
    "INGEST": "NECROBINDER",
    "INITIALIZING_CHARGE": "DEFECT",
    "INSERTAINER": "REGENT",
    "INSIGHT": "COLORLESS",
    "INTIMIDATE": "IRONCLAD",
    "INVEST": "REGENT",
    "IRON_WAVE": "IRONCLAD",
    "JACK": "REGENT",
    "JUGGERNAUT": "IRONCLAD",
    "KNOCKBACK": "IRONCLAD",
    "LEECH": "NECROBINDER",
    "LEFT_HOOK": "REGENT",
    "LEG_SWEEP": "SILENT",
    "LESSON_LEARNED": "DEFECT",
    "LEVELLER": "REGENT",
    "LIGHTNING": "DEFECT",
    "LIMITED": "DEFECT",
    "LOCK_ON": "DEFECT",
    "LOOP": "DEFECT",
    "MAGNETISM": "DEFECT",
    "MALLEABLE": "DEFECT",
    "MASTER_REALITY": "COLORLESS",
    "MAW": "NECROBINDER",
    "MAYHEM": "NECROBINDER",
    "MEGA_DYNAMIC": "IRONCLAD",
    "METALLIC_SCALE": "IRONCLAD",
    "MIND_BLAST": "COLORLESS",
    "MIND_BOLT": "DEFECT",
    "MIRROR": "COLORLESS",
    "MULTICAST": "DEFECT",
    "MULTIPLY": "DEFECT",
    "MUTATION": "REGENT",
    "NEOW_BLESSING": "COLORLESS",
    "NEURO": "DEFECT",
    "NIGHTMARE": "DEFECT",
    "NOBLE": "REGENT",
    "OBLITERATE": "IRONCLAD",
    "OMEGA": "DEFECT",
    "OMNISCIENCE": "COLORLESS",
    "OUTMANEUVER": "REGENT",
    "PALLBEARER": "NECROBINDER",
    "PANACEA": "COLORLESS",
    "PARASITE": "SILENT",
    "PATH_TO_VIOLENCE": "IRONCLAD",
    "PAYDAY": "COLORLESS",
    "PEACE": "REGENT",
    "PELOTON": "DEFECT",
    "PERFECTED_STRIKE": "IRONCLAD",
    "PERFECTLY_FINE": "REGENT",
    "PHANTOM": "SILENT",
    "PHLEGM": "DEFECT",
    "PICK_POCKET": "SILENT",
    "PIERCING_BROW": "DEFECT",
    "PIERCING_WAIL": "SILENT",
    "PILLAGE": "IRONCLAD",
    "POISON": "SILENT",
    "POMMEL_STRIKE": "IRONCLAD",
    "POWER_THROUGH": "IRONCLAD",
    "PRAETORIAN": "REGENT",
    "PRECISION": "SILENT",
    "PREDATOR": "SILENT",
    "PREP_TIME": "COLORLESS",
    "PREPARED": "SILENT",
    "PRIDE": "IRONCLAD",
    "PRIORITY": "DEFECT",
    "PROSTRATE": "IRONCLAD",
    "PROWL": "SILENT",
    "PSIONIC": "DEFECT",
    "PUMMEL": "IRONCLAD",
    "PURE": "DEFECT",
    "QUAKE": "DEFECT",
    "RAIN": "DEFECT",
    "RAMPAGE": "IRONCLAD",
    "REACH": "DEFECT",
    "REACTIVE": "DEFECT",
    "REBOOT": "DEFECT",
    "RECALL": "DEFECT",
    "RECKLESS": "REGENT",
    "RECUR": "DEFECT",
    "RECYCLE": "DEFECT",
    "REDLINE": "DEFECT",
    "REFLECT": "REGENT",
    "REGRET": "COLORLESS",
    "REINFORCE": "DEFECT",
    "RELEASE": "REGENT",
    "REND": "DEFECT",
    "REST": "IRONCLAD",
    "REVENGE": "IRONCLAD",
    "RITUAL": "NECROBINDER",
    "ROUGH": "REGENT",
    "RUTHLESS": "IRONCLAD",
    "SACRIFICE": "NECROBINDER",
    "SAFEGUARD": "REGENT",
    "SANCTITY": "SILENT",
    "SAND": "DEFECT",
    "SCEPTICISM": "DEFECT",
    "SCOOP": "REGENT",
    "SCRAPE": "SILENT",
    "SEARING_BLOW": "IRONCLAD",
    "SECOND_WIND": "IRONCLAD",
    "SECRET_TECHNIQUE": "COLORLESS",
    "SECRET_WEAPON": "COLORLESS",
    "SEDATION": "DEFECT",
    "SEEKER": "REGENT",
    "SEIZE": "NECROBINDER",
    "SELF_REPAIR": "IRONCLAD",
    "SENTINEL": "IRONCLAD",
    "SETUP": "SILENT",
    "SEVER": "NECROBINDER",
    "SHANK": "SILENT",
    "SHARPEN": "IRONCLAD",
    "SHIV": "SILENT",
    "SHOOT": "REGENT",
    "SHORT_STRIDE": "REGENT",
    "SHRUG_IT_OFF": "IRONCLAD",
    "SHUFFLE": "REGENT",
    "SIGHT": "DEFECT",
    "SIGN": "DEFECT",
    "SKIM": "DEFECT",
    "SKIP": "REGENT",
    "SLEEP": "SILENT",
    "SLICE": "REGENT",
    "SLIMED": "COLORLESS",
    "SLOTH": "REGENT",
    "SMITE": "IRONCLAD",
    "SMOKE_BOMB": "SILENT",
    "SNAP": "NECROBINDER",
    "SNIPER": "DEFECT",
    "SOLAR": "IRONCLAD",
    "SONIC_BURST": "DEFECT",
    "SOUL_STRIKE": "NECROBINDER",
    "SPECULATOR": "REGENT",
    "SPIKED_ARMOR": "IRONCLAD",
    "SPIRAL": "DEFECT",
    "SPIT": "DEFECT",
    "SPOT_WEAKNESS": "IRONCLAD",
    "STACK": "REGENT",
    "STAMPEDE": "IRONCLAD",
    "STAR": "DEFECT",
    "STASIS": "DEFECT",
    "STATIC": "DEFECT",
    "STIM": "REGENT",
    "STIM_PACK": "DEFECT",
    "STOCK": "REGENT",
    "STOLE": "SILENT",
    "STORM": "DEFECT",
    "STRIKE": "IRONCLAD",
    "STRIKE_DEFECT": "DEFECT",
    "STRIKE_NECROBINDER": "NECROBINDER",
    "STRIKE_REGENT": "REGENT",
    "STRIKE_SILENT": "SILENT",
    "SUICIDE": "NECROBINDER",
    "SUNDER": "IRONCLAD",
    "SUPLEX": "IRONCLAD",
    "SUPPRESSION": "DEFECT",
    "SURVEILLANCE": "DEFECT",
    "SWEEP": "DEFECT",
    "SWIFT_STRIKE": "SILENT",
    "SWORD": "REGENT",
    "SYNC": "DEFECT",
    "TAKE": "DEFECT",
    "TANK": "IRONCLAD",
    "TEAM_SLAY": "REGENT",
    "TEAR": "DEFECT",
    "TELECAST": "DEFECT",
    "TEMPEST": "DEFECT",
    "TENDRILS": "NECROBINDER",
    "THANATOS": "NECROBINDER",
    "THUNDER": "DEFECT",
    "TINDER": "IRONCLAD",
    "TONE": "DEFECT",
    "TONGUE": "REGENT",
    "TOXIC": "SILENT",
    "TRIP": "SILENT",
    "TURBO": "DEFECT",
    "TWIN_STRIKE": "IRONCLAD",
    "ULTIMATE": "DEFECT",
    "UNCHARGED": "DEFECT",
    "UNDERHAND": "REGENT",
    "UNDERSTANDING": "DEFECT",
    "UNDO": "COLORLESS",
    "UNLOAD": "IRONCLAD",
    "UPPERCUT": "IRONCLAD",
    "UPROOT": "DEFECT",
    "VAULT": "SILENT",
    "VENGEANCE": "REGENT",
    "VENOM": "SILENT",
    "VOLT_EDGE": "DEFECT",
    "VOLTAIC": "DEFECT",
    "WAIL": "DEFECT",
    "WAR_CRY": "IRONCLAD",
    "WARD": "DEFECT",
    "WARP": "DEFECT",
    "WASH": "DEFECT",
    "WATCHER_STANCE": "REGENT",
    "WEAVE": "SILENT",
    "WHEEL": "DEFECT",
    "WHIRLWIND": "IRONCLAD",
    "WISH": "COLORLESS",
    "WONDER": "REGENT",
    "WORLD": "DEFECT",
    "WRATH": "REGENT",
    "WREATH": "REGENT",
    "WRATH_FORM": "REGENT",
    "ZAP": "DEFECT",
    "ZEN": "REGENT",
}


def find_sts2_history_dir():
    if sys.platform == "win32":
        paths = [
            os.path.join(os.environ.get("LOCALAPPDATA", ""), "SlayTheSpire2", "steam"),
        ]
    elif sys.platform == "darwin":
        paths = [
            os.path.expanduser("~/Library/Application Support/SlayTheSpire2/steam/"),
        ]
    else:
        paths = [
            os.path.expanduser("~/.local/share/SlayTheSpire2/steam/"),
        ]

    for base_path in paths:
        if not os.path.exists(base_path):
            continue
        try:
            for entry in os.listdir(base_path):
                profile_path = os.path.join(base_path, entry, "saves", "history")
                if os.path.isdir(profile_path) and glob.glob(
                    os.path.join(profile_path, "*.run")
                ):
                    return profile_path
        except PermissionError:
            continue
    return None


def load_runs(history_dir):
    runs = []
    for f in sorted(glob.glob(os.path.join(history_dir, "*.run"))):
        try:
            with open(f, "r") as file:
                runs.append(json.load(file))
        except:
            continue
    return runs


def get_card_data(runs):
    pick_by_class = defaultdict(
        lambda: defaultdict(lambda: {"offered": 0, "picked": 0})
    )
    win_by_class = defaultdict(lambda: defaultdict(lambda: {"runs": 0, "wins": 0}))

    for run in runs:
        is_win = run.get("win", False)
        players = run.get("players", [])
        if not players:
            continue
        char = players[0].get("character", "Unknown").replace("CHARACTER.", "")
        if char not in CLASS_COLORS:
            continue

        cards_in_run = set()

        for act in run.get("map_point_history", []):
            for point in act:
                for stat in point.get("player_stats", []):
                    for choice in stat.get("card_choices", []):
                        card_id = (
                            choice.get("card", {})
                            .get("id", "Unknown")
                            .replace("CARD.", "")
                        )
                        was_picked = choice.get("was_picked", False)

                        pick_by_class[char][card_id]["offered"] += 1
                        if was_picked:
                            pick_by_class[char][card_id]["picked"] += 1
                            cards_in_run.add(card_id)

        for cid in cards_in_run:
            win_by_class[char][cid]["runs"] += 1
            if is_win:
                win_by_class[char][cid]["wins"] += 1

    return pick_by_class, win_by_class


def create_excel(pick_by_class, win_by_class, output_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    for class_name in CLASS_COLORS.keys():
        ws = wb.create_sheet(title=class_name)
        header_fill = PatternFill(
            start_color=CLASS_COLORS[class_name],
            end_color=CLASS_COLORS[class_name],
            fill_type="solid",
        )
        headers = [
            "Card",
            "Original Class",
            "Times Offered",
            "Times Picked",
            "Pick Rate %",
            "Runs with Card",
            "Wins with Card",
            "Win Rate %",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        row = 2
        class_cards = list(pick_by_class[class_name].items())
        class_cards.sort(key=lambda x: x[1]["picked"], reverse=True)

        for card_id, pick_stats in class_cards:
            win_stats = win_by_class[class_name].get(card_id, {"runs": 0, "wins": 0})
            original_class = CARD_CLASSIFICATIONS.get(card_id, "COLORLESS")
            pick_rate = (
                (pick_stats["picked"] / pick_stats["offered"] * 100)
                if pick_stats["offered"] > 0
                else 0
            )
            win_rate = (
                (win_stats["wins"] / win_stats["runs"] * 100)
                if win_stats["runs"] > 0
                else 0
            )

            ws.cell(row=row, column=1, value=card_id)
            ws.cell(row=row, column=2, value=original_class)
            ws.cell(row=row, column=3, value=pick_stats["offered"])
            ws.cell(row=row, column=4, value=pick_stats["picked"])
            ws.cell(row=row, column=5, value=round(pick_rate, 1))
            ws.cell(row=row, column=6, value=win_stats["runs"])
            ws.cell(row=row, column=7, value=win_stats["wins"])
            ws.cell(row=row, column=8, value=round(win_rate, 1))
            row += 1

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        for col in range(3, 9):
            ws.column_dimensions[get_column_letter(col)].width = 14

    wb.save(output_path)
    return sum(len(cards) for cards in pick_by_class.values())


class STSCardViewer(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("STS2 Card Statistics")
        self.geometry("1100x650")
        self.history_dir = None
        self.card_data = {}
        self.all_data = []
        self.current_class = None

        self.create_widgets()
        self.find_and_load_data()

    def create_widgets(self):
        top_frame = ttk.Frame(self)
        top_frame.pack(fill="x", padx=5, pady=5)

        ttk.Label(top_frame, text="Select STS2 Save Location:").pack(
            side="left", padx=5
        )

        self.path_label = ttk.Label(top_frame, text="Not selected", foreground="gray")
        self.path_label.pack(side="left", padx=5)

        ttk.Button(top_frame, text="Browse...", command=self.browse_folder).pack(
            side="left", padx=5
        )
        ttk.Button(top_frame, text="Generate Data", command=self.generate_data).pack(
            side="left", padx=5
        )
        ttk.Button(top_frame, text="Help", command=self.show_help).pack(
            side="right", padx=5
        )

        toolbar = ttk.Frame(self)
        toolbar.pack(fill="x", padx=5, pady=5)

        ttk.Label(toolbar, text="Class:").pack(side="left", padx=5)
        self.class_var = tk.StringVar(value="IRONCLAD")
        for class_name in CLASS_COLORS.keys():
            rb = ttk.Radiobutton(
                toolbar,
                text=class_name,
                value=class_name,
                variable=self.class_var,
                command=self.on_class_change,
            )
            rb.pack(side="left", padx=2)

        ttk.Label(toolbar, text="Search:").pack(side="left", padx=(20, 5))
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(toolbar, textvariable=self.search_var, width=20)
        self.search_entry.pack(side="left")
        self.search_entry.bind("<KeyRelease>", lambda e: self.filter_data())

        ttk.Button(toolbar, text="Refresh", command=self.refresh).pack(
            side="right", padx=5
        )

        main_frame = ttk.Frame(self)
        main_frame.pack(fill="both", expand=True, padx=5, pady=5)

        columns = ("Card", "Orig", "Offered", "Picked", "Pick%", "Runs", "Wins", "Win%")
        self.tree = ttk.Treeview(main_frame, columns=columns, show="headings")

        self.tree.heading("Card", text="Card", command=lambda: self.sort_by("card"))
        self.tree.heading(
            "Orig", text="Orig Class", command=lambda: self.sort_by("original_class")
        )
        self.tree.heading(
            "Offered", text="Times Offered", command=lambda: self.sort_by("offered")
        )
        self.tree.heading(
            "Picked", text="Times Picked", command=lambda: self.sort_by("picked")
        )
        self.tree.heading(
            "Pick%", text="Pick Rate %", command=lambda: self.sort_by("pick_rate")
        )
        self.tree.heading(
            "Runs", text="Runs with Card", command=lambda: self.sort_by("runs")
        )
        self.tree.heading(
            "Wins", text="Wins with Card", command=lambda: self.sort_by("wins")
        )
        self.tree.heading(
            "Win%", text="Win Rate %", command=lambda: self.sort_by("win_rate")
        )

        self.tree.column("Card", width=180)
        self.tree.column("Orig", width=80, anchor="center")
        for col in columns[2:]:
            self.tree.column(col, width=70, anchor="e")

        scrollbar = ttk.Scrollbar(
            main_frame, orient="vertical", command=self.tree.yview
        )
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.tree.tag_configure("oddrow", background="#f0f0f0")
        self.tree.tag_configure("evenrow", background="#ffffff")

        self.status_label = ttk.Label(self, text="", relief="sunken", anchor="w")
        self.status_label.pack(fill="x", padx=5, pady=2)

    def find_and_load_data(self):
        self.history_dir = find_sts2_history_dir()
        if self.history_dir:
            self.path_label.config(text=self.history_dir, foreground="black")
            self.load_excel_data()
        else:
            self.path_label.config(
                text="Not found - click Browse to select", foreground="red"
            )

    def browse_folder(self):
        folder = filedialog.askdirectory(title="Select STS2 save folder")
        if folder:
            history_path = os.path.join(folder, "saves", "history")
            if os.path.exists(history_path) and glob.glob(
                os.path.join(history_path, "*.run")
            ):
                self.history_dir = history_path
                self.path_label.config(text=history_path, foreground="black")
                self.generate_data()
            else:
                self.history_dir = folder
                self.path_label.config(text=folder, foreground="black")
                self.generate_data()

    def generate_data(self):
        if not self.history_dir:
            messagebox.showwarning("Warning", "Please select a STS2 save folder first.")
            return

        self.status_label.config(text="Loading runs...")
        self.update()

        runs = load_runs(self.history_dir)
        if not runs:
            messagebox.showwarning(
                "No Data", "No run files found in the selected folder."
            )
            self.status_label.config(text="No runs found")
            return

        self.status_label.config(text=f"Processing {len(runs)} runs...")
        self.update()

        pick_by_class, win_by_class = get_card_data(runs)
        total = create_excel(pick_by_class, win_by_class, EXCEL_FILE)

        self.status_label.config(text=f"Generated {total} cards from {len(runs)} runs")
        self.load_excel_data()

    def load_excel_data(self):
        if not os.path.exists(EXCEL_FILE):
            return

        self.card_data = {}
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            for class_name in wb.sheetnames:
                ws = wb[class_name]
                self.card_data[class_name] = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        self.card_data[class_name].append(
                            {
                                "card": row[0],
                                "original_class": row[1],
                                "offered": row[2],
                                "picked": row[3],
                                "pick_rate": row[4],
                                "runs": row[5],
                                "wins": row[6],
                                "win_rate": row[7],
                            }
                        )
            self.load_class("IRONCLAD")
            runs = load_runs(self.history_dir) if self.history_dir else []
            self.status_label.config(
                text=f"Loaded {sum(len(c) for c in self.card_data.values())} cards from {len(runs)} runs"
            )
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load data: {e}")

    def load_class(self, class_name):
        self.current_class = class_name
        self.all_data = self.card_data.get(class_name, [])
        self.filter_data()

    def filter_data(self):
        search = self.search_var.get().lower()
        filtered = self.all_data
        if search:
            filtered = [c for c in self.all_data if search in c["card"].lower()]

        for row in self.tree.get_children():
            self.tree.delete(row)

        for i, card in enumerate(filtered):
            tags = ("oddrow",) if i % 2 == 0 else ("evenrow",)
            self.tree.insert(
                "",
                "end",
                values=(
                    card["card"],
                    card["original_class"],
                    card["offered"],
                    card["picked"],
                    card["pick_rate"],
                    card["runs"],
                    card["wins"],
                    card["win_rate"],
                ),
                tags=tags,
            )

    def sort_by(self, key):
        reverse = getattr(self, "_sort_reverse", False)
        self._sort_reverse = not reverse

        def get_val(card):
            val = card.get(key)
            if isinstance(val, str):
                return val.lower()
            return val if val else 0

        self.all_data.sort(key=get_val, reverse=reverse)
        self.filter_data()

    def on_class_change(self):
        self.load_class(self.class_var.get())

    def refresh(self):
        self.find_and_load_data()

    def show_help(self):
        help_text = """How to find your STS2 save folder:

WINDOWS:
%LOCALAPPDATA%\\SlayTheSpire2\\steam\\<profile>\\saves\\history

To open: Press Win+R, type %LOCALAPPDATA%, press Enter
Then navigate to SlayTheSpire2 > steam > (profile folder) > saves > history

MAC:
~/Library/Application Support/SlayTheSpire2/steam/<profile>/saves/history

To open: Press Cmd+Shift+G in Finder, paste the path above

LINUX:
~/.local/share/SlayTheSpire2/steam/<profile>/saves/history

To open: Press Ctrl+L in file manager, paste ~/.local/share/SlayTheSpire2/

NOTE: <profile> is usually a number like 76561198054269638

If auto-detection fails, click "Browse" and select the folder 
containing your .run files (usually the "history" folder)."""
        messagebox.showinfo("Help - Finding STS2 Save Files", help_text)


if __name__ == "__main__":
    app = STSCardViewer()
    app.mainloop()
