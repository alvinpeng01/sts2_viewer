#!/usr/bin/env python3
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
import os
import sys
import glob
import json
from collections import defaultdict, Counter
from data.card_classifications import CARD_CLASSIFICATIONS
from data.relic_info import RELIC_INFO
from data.event_info import EVENT_INFO
from data.event_choices import EVENT_CHOICES

if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

EXCEL_FILE = os.path.join(os.getcwd(), "sts2_cards.xlsx")

CLASS_COLORS = {
    "IRONCLAD": "FF8B8B",
    "SILENT": "8BFF8B",
    "DEFECT": "8B8BFF",
    "NECROBINDER": "FF8BFF",
    "REGENT": "FFFF8B",
}


def find_sts2_history_dir():
    if sys.platform == "win32":
        paths = [
            os.path.join(os.environ.get("APPDATA", ""), "SlayTheSpire2", "steam"),
            os.path.join(os.environ.get("LOCALAPPDATA", ""), "SlayTheSpire2", "steam"),
        ]
    elif sys.platform == "darwin":
        paths = [
            os.path.expanduser("~/Library/Application Support/SlayTheSpire2/steam/")
        ]
    else:
        paths = [os.path.expanduser("~/.local/share/SlayTheSpire2/steam/")]

    for base_path in paths:
        if not os.path.exists(base_path):
            continue
        try:
            for entry in os.listdir(base_path):
                profile_base = os.path.join(base_path, entry)
                for subpath in ["profile1/saves/history", "saves/history"]:
                    profile_path = os.path.join(profile_base, subpath)
                    if os.path.isdir(profile_path) and glob.glob(
                        os.path.join(profile_path, "*.run")
                    ):
                        return profile_path
        except PermissionError:
            continue
    return None


def load_runs(history_dir):
    runs = []
    for f in sorted(glob.glob(os.path.join(history_dir, "*.run"))):
        try:
            with open(f, "r") as file:
                runs.append(json.load(file))
        except:
            continue
    return runs


def get_card_data(runs):
    pick_by_class = defaultdict(
        lambda: defaultdict(lambda: {"offered": 0, "picked": 0})
    )
    win_by_class = defaultdict(lambda: defaultdict(lambda: {"runs": 0, "wins": 0}))

    for run in runs:
        is_win = run.get("win", False)
        players = run.get("players", [])
        if not players:
            continue
        char = players[0].get("character", "Unknown").replace("CHARACTER.", "")
        if char not in CLASS_COLORS:
            continue

        cards_in_run = set()
        for act in run.get("map_point_history", []):
            for point in act:
                for stat in point.get("player_stats", []):
                    for choice in stat.get("card_choices", []):
                        card_id = (
                            choice.get("card", {})
                            .get("id", "Unknown")
                            .replace("CARD.", "")
                        )
                        was_picked = choice.get("was_picked", False)
                        pick_by_class[char][card_id]["offered"] += 1
                        if was_picked:
                            pick_by_class[char][card_id]["picked"] += 1
                            cards_in_run.add(card_id)

        for cid in cards_in_run:
            win_by_class[char][cid]["runs"] += 1
            if is_win:
                win_by_class[char][cid]["wins"] += 1

    return pick_by_class, win_by_class


def get_relic_data(runs):
    relic_stats = defaultdict(
        lambda: {
            "total": 0,
            "wins": 0,
            "by_class": defaultdict(lambda: {"total": 0, "wins": 0}),
        }
    )
    for run in runs:
        is_win = run.get("win", False)
        players = run.get("players", [])
        if not players:
            continue
        char = players[0].get("character", "Unknown").replace("CHARACTER.", "")
        if char not in CLASS_COLORS:
            continue
        for relic in players[0].get("relics", []):
            rid = relic.get("id", "Unknown").replace("RELIC.", "")
            relic_stats[rid]["total"] += 1
            relic_stats[rid]["by_class"][char]["total"] += 1
            if is_win:
                relic_stats[rid]["wins"] += 1
                relic_stats[rid]["by_class"][char]["wins"] += 1
    return relic_stats


def get_ancient_relic_data(runs):
    """Track ancient relic choices - which relics were offered and picked at ancient events"""
    ancient_stats = defaultdict(
        lambda: defaultdict(
            lambda: {
                "offered": 0,
                "picked": 0,
                "wins": 0,
                "by_class": defaultdict(lambda: {"offered": 0, "picked": 0, "wins": 0}),
            }
        )
    )
    # Filter out non-relic values
    invalid_keys = {"IRONCLAD", "SILENT", "DEFECT", "NECROBINDER", "REGENT", ""}
    for run in runs:
        is_win = run.get("win", False)
        players = run.get("players", [])
        if not players:
            continue
        char = players[0].get("character", "Unknown").replace("CHARACTER.", "")
        if char not in CLASS_COLORS:
            continue
        for act in run.get("map_point_history", []):
            for point in act:
                if point.get("map_point_type") != "ancient":
                    continue
                rooms = point.get("rooms", [])
                if not rooms:
                    continue
                event_id = rooms[0].get("model_id", "").replace("EVENT.", "")
                for stat in point.get("player_stats", []):
                    for ancient_choice in stat.get("ancient_choice", []):
                        rid = ancient_choice.get("TextKey", "").replace("RELIC.", "")
                        if rid in invalid_keys:
                            continue
                        was_chosen = ancient_choice.get("was_chosen", False)
                        ancient_stats[event_id][rid]["offered"] += 1
                        ancient_stats[event_id][rid]["by_class"][char]["offered"] += 1
                        if was_chosen:
                            ancient_stats[event_id][rid]["picked"] += 1
                            ancient_stats[event_id][rid]["by_class"][char][
                                "picked"
                            ] += 1
                        if was_chosen and is_win:
                            ancient_stats[event_id][rid]["wins"] += 1
                            ancient_stats[event_id][rid]["by_class"][char]["wins"] += 1
    return ancient_stats


def get_encounter_data(runs):
    """Track damage taken from each encounter"""
    encounter_stats = defaultdict(
        lambda: {
            "count": 0,
            "total_damage": 0,
            "wins": 0,
            "acts": defaultdict(int),
            "enemies": Counter(),
        }
    )
    for run in runs:
        is_win = run.get("win", False)
        for act_idx, act in enumerate(run.get("map_point_history", [])):
            for point in act:
                rooms = point.get("rooms", [])
                if not rooms:
                    continue
                room = rooms[0]
                room_id = room.get("model_id", "")
                if "ENCOUNTER" not in room_id:
                    continue
                encounter_id = room_id.replace("ENCOUNTER.", "")
                monster_ids = room.get("monster_ids", [])
                for stat in point.get("player_stats", []):
                    damage = stat.get("damage_taken", 0)
                    encounter_stats[encounter_id]["count"] += 1
                    encounter_stats[encounter_id]["total_damage"] += damage
                    encounter_stats[encounter_id]["acts"][act_idx + 1] += 1
                    for m in monster_ids:
                        m_name = m.replace("MONSTER.", "")
                        encounter_stats[encounter_id]["enemies"][m_name] += 1
                    if is_win:
                        encounter_stats[encounter_id]["wins"] += 1
    return encounter_stats


def get_event_data(runs):
    event_stats = defaultdict(
        lambda: defaultdict(
            lambda: {
                "total": 0,
                "wins": 0,
                "encounters": 0,
                "is_ancient": False,
                "by_class": defaultdict(
                    lambda: {"total": 0, "wins": 0, "encounters": 0}
                ),
            }
        )
    )
    for run in runs:
        is_win = run.get("win", False)
        players = run.get("players", [])
        if not players:
            continue
        char = players[0].get("character", "Unknown").replace("CHARACTER.", "")
        if char not in CLASS_COLORS:
            continue
        for act in run.get("map_point_history", []):
            for point in act:
                rooms = point.get("rooms", [])
                if not rooms:
                    continue
                room = rooms[0]
                if room.get("room_type") != "event":
                    continue
                event_id = room.get("model_id", "").replace("EVENT.", "")
                is_ancient = point.get("map_point_type") == "ancient"
                event_stats[event_id]["_meta"]["encounters"] += 1
                event_stats[event_id]["_meta"]["is_ancient"] = is_ancient
                event_stats[event_id]["_meta"]["by_class"][char]["encounters"] += 1
                if is_win:
                    event_stats[event_id]["_meta"]["wins"] += 1
                    event_stats[event_id]["_meta"]["by_class"][char]["wins"] += 1
                for stat in point.get("player_stats", []):
                    for choice in stat.get("event_choices", []):
                        title_key = choice.get("title", {}).get("key", "")
                        if "options." in title_key:
                            choice_id = title_key.split("options.")[-1].split(".")[0]
                        else:
                            choice_id = "DEFAULT"
                        event_stats[event_id][choice_id]["total"] += 1
                        event_stats[event_id][choice_id]["by_class"][char]["total"] += 1
                        if is_win:
                            event_stats[event_id][choice_id]["wins"] += 1
                            event_stats[event_id][choice_id]["by_class"][char][
                                "wins"
                            ] += 1
    return event_stats


def create_excel(
    pick_by_class,
    win_by_class,
    relic_stats,
    event_stats,
    ancient_stats,
    encounter_stats,
    output_path,
):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for class_name in CLASS_COLORS.keys():
        ws = wb.create_sheet(title=class_name)
        header_fill = PatternFill(
            start_color="FF" + CLASS_COLORS[class_name],
            end_color="FF" + CLASS_COLORS[class_name],
            fill_type="solid",
        )
        headers = [
            "Card",
            "Original Class",
            "Times Offered",
            "Times Picked",
            "Pick Rate %",
            "Runs with Card",
            "Wins with Card",
            "Win Rate %",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        row = 2
        class_cards = list(pick_by_class[class_name].items())
        class_cards.sort(key=lambda x: x[1]["picked"], reverse=True)

        for card_id, pick_stats in class_cards:
            win_stats = win_by_class[class_name].get(card_id, {"runs": 0, "wins": 0})
            original_class = CARD_CLASSIFICATIONS.get(card_id, "COLORLESS")
            pick_rate = (
                (pick_stats["picked"] / pick_stats["offered"] * 100)
                if pick_stats["offered"] > 0
                else 0
            )
            win_rate = (
                (win_stats["wins"] / win_stats["runs"] * 100)
                if win_stats["runs"] > 0
                else 0
            )

            ws.cell(row=row, column=1, value=card_id)
            ws.cell(row=row, column=2, value=original_class)
            ws.cell(row=row, column=3, value=pick_stats["offered"])
            ws.cell(row=row, column=4, value=pick_stats["picked"])
            ws.cell(row=row, column=5, value=round(pick_rate, 1))
            ws.cell(row=row, column=6, value=win_stats["runs"])
            ws.cell(row=row, column=7, value=win_stats["wins"])
            ws.cell(row=row, column=8, value=round(win_rate, 1))
            row += 1

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        for col in range(3, 9):
            ws.column_dimensions[get_column_letter(col)].width = 14

    # Add Relics sheet
    ws = wb.create_sheet(title="Relics")
    header_fill = PatternFill(
        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
    )
    headers = ["ID", "Name", "Runs", "Wins", "Win%", "", "Description"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Add class headers
    for i, class_name in enumerate(CLASS_COLORS.keys()):
        ws.cell(row=1, column=8 + i * 3, value=class_name)
        ws.cell(row=1, column=8 + i * 3 + 1, value="W")
        ws.cell(row=1, column=8 + i * 3 + 2, value="%")

    row = 2
    relics_list = [(rid, stats) for rid, stats in relic_stats.items()]
    relics_list.sort(key=lambda x: x[1]["total"], reverse=True)

    for relic_id, stats in relics_list:
        win_rate = (stats["wins"] / stats["total"] * 100) if stats["total"] > 0 else 0
        relic_info = RELIC_INFO.get(
            relic_id, {"name": relic_id, "description": "Unknown"}
        )

        ws.cell(row=row, column=1, value=relic_id)
        ws.cell(row=row, column=2, value=relic_info.get("name", relic_id))
        ws.cell(row=row, column=3, value=stats["total"])
        ws.cell(row=row, column=4, value=stats["wins"])
        ws.cell(row=row, column=5, value=round(win_rate, 1))
        ws.cell(row=row, column=6, value="")
        ws.cell(row=row, column=7, value=relic_info.get("description", ""))

        for i, class_name in enumerate(CLASS_COLORS.keys()):
            class_stats = stats["by_class"].get(class_name, {"total": 0, "wins": 0})
            class_win_rate = (
                (class_stats["wins"] / class_stats["total"] * 100)
                if class_stats["total"] > 0
                else 0
            )
            ws.cell(row=row, column=8 + i * 3, value=class_stats["total"])
            ws.cell(row=row, column=8 + i * 3 + 1, value=class_stats["wins"])
            ws.cell(
                row=row,
                column=8 + i * 3 + 2,
                value=round(class_win_rate, 1) if class_stats["total"] > 0 else "-",
            )
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 2
    ws.column_dimensions["G"].width = 45
    for i in range(len(CLASS_COLORS)):
        ws.column_dimensions[get_column_letter(8 + i * 3)].width = 8
        ws.column_dimensions[get_column_letter(8 + i * 3 + 1)].width = 5
        ws.column_dimensions[get_column_letter(8 + i * 3 + 2)].width = 5

    # Add Events sheet
    ws = wb.create_sheet(title="Events")
    header_fill = PatternFill(
        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
    )
    headers = ["Event", "Choice", "Encounters", "Picked", "Pick%", "Win%", "Wins"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Add class headers
    for i, class_name in enumerate(CLASS_COLORS.keys()):
        ws.cell(row=1, column=8 + i * 3, value=class_name)
        ws.cell(row=1, column=8 + i * 3 + 1, value="W")
        ws.cell(row=1, column=8 + i * 3 + 2, value="%")

    ws.cell(row=1, column=23, value="Description")
    ws.cell(row=1, column=23).fill = header_fill
    ws.cell(row=1, column=23).font = Font(bold=True)

    row = 2

    # Separate ancient and regular events
    ancient_events = []
    regular_events = []

    for event_id, choices in event_stats.items():
        if event_id == "_meta":
            continue
        meta = event_stats[event_id].get(
            "_meta", {"encounters": 0, "wins": 0, "is_ancient": False}
        )
        is_ancient = meta.get("is_ancient", False)
        for choice, stats in choices.items():
            if choice != "_meta":
                ev_data = (event_id, choice, stats, meta)
                if is_ancient:
                    ancient_events.append(ev_data)
                else:
                    regular_events.append(ev_data)

    # Sort by encounters
    ancient_events.sort(key=lambda x: x[3]["encounters"], reverse=True)
    regular_events.sort(key=lambda x: x[3]["encounters"], reverse=True)

    def write_events(start_row, events_list, is_ancient_section=False):
        r = start_row
        if is_ancient_section and events_list:
            # Ancient section header
            ws.cell(row=r, column=1, value="--- ANCIENT EVENTS ---")
            ws.cell(row=r, column=1).font = Font(bold=True, color="000080")
            r += 1

        for event_id, choice, stats, meta in events_list:
            win_rate = (
                (stats["wins"] / stats["total"] * 100) if stats["total"] > 0 else 0
            )
            pick_rate = (
                (stats["total"] / meta["encounters"] * 100)
                if meta["encounters"] > 0
                else 0
            )
            event_info = EVENT_INFO.get(
                event_id, {"name": event_id, "description": "Unknown"}
            )

            # Get choice description
            choice_choices = EVENT_CHOICES.get(event_id, {})
            choice_desc = choice_choices.get(choice, choice)
            if choice == "DEFAULT":
                choice_desc = event_info.get("description", "Default")[:50]

            ws.cell(row=r, column=1, value=event_id)
            ws.cell(row=r, column=2, value=choice_desc)
            ws.cell(row=r, column=3, value=meta["encounters"])
            ws.cell(row=r, column=4, value=stats["total"])
            ws.cell(row=r, column=5, value=round(pick_rate, 1))
            ws.cell(row=r, column=6, value=round(win_rate, 1))
            ws.cell(row=r, column=7, value=stats["wins"])

            for i, class_name in enumerate(CLASS_COLORS.keys()):
                class_stats = stats["by_class"].get(class_name, {"total": 0, "wins": 0})
                class_meta = meta["by_class"].get(
                    class_name, {"encounters": 0, "wins": 0}
                )
                class_pick_rate = (
                    (class_stats["total"] / class_meta["encounters"] * 100)
                    if class_meta["encounters"] > 0
                    else 0
                )
                ws.cell(
                    row=r,
                    column=8 + i * 3,
                    value=round(class_pick_rate, 0)
                    if class_stats["total"] > 0
                    else "-",
                )
                ws.cell(row=r, column=8 + i * 3 + 1, value=class_stats["wins"])
                ws.cell(row=r, column=8 + i * 3 + 2, value="-")

            # Description at the end with highlighting
            desc_cell = ws.cell(
                row=r, column=23, value=event_info.get("description", "")
            )
            desc_cell.fill = PatternFill(
                start_color="FFFACD", end_color="FFFACD", fill_type="solid"
            )
            r += 1

        return r

    # Write regular events only (no ancient events)
    row = write_events(row, regular_events)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 6
    for i in range(len(CLASS_COLORS)):
        ws.column_dimensions[get_column_letter(8 + i * 3)].width = 7
        ws.column_dimensions[get_column_letter(8 + i * 3 + 1)].width = 5
        ws.column_dimensions[get_column_letter(8 + i * 3 + 2)].width = 5
    ws.column_dimensions["W"].width = 5
    ws.column_dimensions["X"].width = 5
    ws.column_dimensions["Y"].width = 50

    # Add Ancient Relics sheet
    ws = wb.create_sheet(title="Ancient Relics")
    header_fill = PatternFill(
        start_color="FFD700", end_color="FFD700", fill_type="solid"
    )
    headers = ["Event", "Relic", "Offered", "Picked", "Pick%", "Wins", "Win%"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Add class headers
    for i, class_name in enumerate(CLASS_COLORS.keys()):
        ws.cell(row=1, column=8 + i * 3, value=class_name)
        ws.cell(row=1, column=8 + i * 3 + 1, value="Off")
        ws.cell(row=1, column=8 + i * 3 + 2, value="Pick%")

    ws.cell(row=1, column=23, value="Relic Description")
    ws.cell(row=1, column=23).fill = header_fill
    ws.cell(row=1, column=23).font = Font(bold=True)

    row = 2
    ancient_list = []
    for event_id, relics in ancient_stats.items():
        for relic_id, stats in relics.items():
            ancient_list.append((event_id, relic_id, stats))
    ancient_list.sort(key=lambda x: x[2]["offered"], reverse=True)

    for event_id, relic_id, stats in ancient_list:
        pick_rate = (
            (stats["picked"] / stats["offered"] * 100) if stats["offered"] > 0 else 0
        )
        win_rate = (stats["wins"] / stats["picked"] * 100) if stats["picked"] > 0 else 0
        relic_info = RELIC_INFO.get(
            relic_id, {"name": relic_id, "description": "Unknown"}
        )

        ws.cell(row=row, column=1, value=event_id)
        ws.cell(row=row, column=2, value=relic_info.get("name", relic_id))
        ws.cell(row=row, column=3, value=stats["offered"])
        ws.cell(row=row, column=4, value=stats["picked"])
        ws.cell(row=row, column=5, value=round(pick_rate, 1))
        ws.cell(row=row, column=6, value=stats["wins"])
        ws.cell(row=row, column=7, value=round(win_rate, 1))

        for i, class_name in enumerate(CLASS_COLORS.keys()):
            class_stats = stats["by_class"].get(class_name, {"offered": 0, "picked": 0})
            class_pick_rate = (
                (class_stats["picked"] / class_stats["offered"] * 100)
                if class_stats["offered"] > 0
                else 0
            )
            ws.cell(row=row, column=8 + i * 3, value=class_stats["offered"])
            ws.cell(row=row, column=8 + i * 3 + 1, value=class_stats["picked"])
            ws.cell(
                row=row,
                column=8 + i * 3 + 2,
                value=round(class_pick_rate, 0) if class_stats["offered"] > 0 else "-",
            )

        desc_cell = ws.cell(row=row, column=23, value=relic_info.get("description", ""))
        desc_cell.fill = PatternFill(
            start_color="FFFACD", end_color="FFFACD", fill_type="solid"
        )
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 8
    for i in range(len(CLASS_COLORS)):
        ws.column_dimensions[get_column_letter(8 + i * 3)].width = 6
        ws.column_dimensions[get_column_letter(8 + i * 3 + 1)].width = 6
        ws.column_dimensions[get_column_letter(8 + i * 3 + 2)].width = 7
        ws.column_dimensions["W"].width = 5
        ws.column_dimensions["X"].width = 5
        ws.column_dimensions["Y"].width = 50

    # Add Encounters sheet
    ws = wb.create_sheet(title="Encounters")
    header_fill = PatternFill(
        start_color="CC6666", end_color="CC6666", fill_type="solid"
    )
    headers = [
        "Encounter",
        "Act#",
        "Enemy Name",
        "Times",
        "Total Damage",
        "Avg Damage",
        "Wins",
        "Win%",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    row = 2
    encounters_list = [(eid, stats) for eid, stats in encounter_stats.items()]
    encounters_list.sort(key=lambda x: x[1]["count"], reverse=True)

    for encounter_id, stats in encounters_list:
        avg_damage = stats["total_damage"] / stats["count"] if stats["count"] > 0 else 0
        win_rate = (stats["wins"] / stats["count"] * 100) if stats["count"] > 0 else 0

        most_common_act = (
            max(stats["acts"].items(), key=lambda x: x[1])[0] if stats["acts"] else "-"
        )
        enemy_list = ", ".join(stats["enemies"].keys()) if stats["enemies"] else "-"

        ws.cell(row=row, column=1, value=encounter_id)
        ws.cell(row=row, column=2, value=most_common_act)
        ws.cell(row=row, column=3, value=enemy_list)
        ws.cell(row=row, column=4, value=stats["count"])
        ws.cell(row=row, column=5, value=stats["total_damage"])
        ws.cell(row=row, column=6, value=round(avg_damage, 1))
        ws.cell(row=row, column=7, value=stats["wins"])
        ws.cell(row=row, column=8, value=round(win_rate, 1))
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 10

    wb.save(output_path)
    return sum(len(cards) for cards in pick_by_class.values())


class STSCardViewer(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("STS2 Statistics Viewer")
        self.geometry("1200x700")
        self.history_dir = None
        self.card_data = {}
        self.relic_data = []
        self.event_data = []
        self.ancient_data = []
        self.encounters_data = []
        self.all_data = []
        self.current_class = None

        self.create_widgets()
        self.find_and_load_data()

    def create_widgets(self):
        top_frame = ttk.Frame(self)
        top_frame.pack(fill="x", padx=5, pady=5)

        ttk.Label(top_frame, text="STS2 Save Location:").pack(side="left", padx=5)
        self.path_label = ttk.Label(top_frame, text="Not selected", foreground="gray")
        self.path_label.pack(side="left", padx=5)
        ttk.Button(top_frame, text="Browse...", command=self.browse_folder).pack(
            side="left", padx=5
        )
        ttk.Button(top_frame, text="Generate Data", command=self.generate_data).pack(
            side="left", padx=5
        )
        ttk.Button(top_frame, text="Help", command=self.show_help).pack(
            side="right", padx=5
        )

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=5, pady=5)

        self.cards_frame = ttk.Frame(self.notebook)
        self.relics_frame = ttk.Frame(self.notebook)
        self.events_frame = ttk.Frame(self.notebook)
        self.ancient_frame = ttk.Frame(self.notebook)
        self.encounters_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.cards_frame, text="Cards")
        self.notebook.add(self.relics_frame, text="Relics")
        self.notebook.add(self.events_frame, text="Events")
        self.notebook.add(self.ancient_frame, text="Ancient Relics")
        self.notebook.add(self.encounters_frame, text="Encounters")

        self.create_cards_tab()
        self.create_relics_tab()
        self.create_events_tab()
        self.create_ancient_tab()
        self.create_encounters_tab()

        self.status_label = ttk.Label(self, text="", relief="sunken", anchor="w")
        self.status_label.pack(fill="x", padx=5, pady=2)

    def create_cards_tab(self):
        toolbar = ttk.Frame(self.cards_frame)
        toolbar.pack(fill="x", padx=5, pady=5)

        ttk.Label(toolbar, text="Class:").pack(side="left", padx=5)
        self.class_var = tk.StringVar(value="IRONCLAD")
        for class_name in CLASS_COLORS.keys():
            rb = ttk.Radiobutton(
                toolbar,
                text=class_name,
                value=class_name,
                variable=self.class_var,
                command=self.on_class_change,
            )
            rb.pack(side="left", padx=2)

        ttk.Label(toolbar, text="Search:").pack(side="left", padx=(20, 5))
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(toolbar, textvariable=self.search_var, width=20)
        self.search_entry.pack(side="left")
        self.search_entry.bind("<KeyRelease>", lambda e: self.filter_data())
        ttk.Button(toolbar, text="Refresh", command=self.refresh).pack(
            side="right", padx=5
        )

        main_frame = ttk.Frame(self.cards_frame)
        main_frame.pack(fill="both", expand=True, padx=5, pady=5)

        columns = ("Card", "Orig", "Offered", "Picked", "Pick%", "Runs", "Wins", "Win%")
        self.tree = ttk.Treeview(main_frame, columns=columns, show="headings")

        self.tree.heading("Card", text="Card", command=lambda: self.sort_by("card"))
        self.tree.heading(
            "Orig", text="Orig Class", command=lambda: self.sort_by("original_class")
        )
        self.tree.heading(
            "Offered", text="Times Offered", command=lambda: self.sort_by("offered")
        )
        self.tree.heading(
            "Picked", text="Times Picked", command=lambda: self.sort_by("picked")
        )
        self.tree.heading(
            "Pick%", text="Pick Rate %", command=lambda: self.sort_by("pick_rate")
        )
        self.tree.heading(
            "Runs", text="Runs with Card", command=lambda: self.sort_by("runs")
        )
        self.tree.heading(
            "Wins", text="Wins with Card", command=lambda: self.sort_by("wins")
        )
        self.tree.heading(
            "Win%", text="Win Rate %", command=lambda: self.sort_by("win_rate")
        )

        self.tree.column("Card", width=180)
        self.tree.column("Orig", width=80, anchor="center")
        for col in columns[2:]:
            self.tree.column(col, width=70, anchor="e")

        scrollbar = ttk.Scrollbar(
            main_frame, orient="vertical", command=self.tree.yview
        )
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        self.tree.tag_configure("oddrow", background="#f0f0f0")
        self.tree.tag_configure("evenrow", background="#ffffff")

    def create_relics_tab(self):
        toolbar = ttk.Frame(self.relics_frame)
        toolbar.pack(fill="x", padx=5, pady=5)

        ttk.Label(toolbar, text="Search:").pack(side="left", padx=5)
        self.relic_search_var = tk.StringVar()
        self.relic_search_entry = ttk.Entry(
            toolbar, textvariable=self.relic_search_var, width=20
        )
        self.relic_search_entry.pack(side="left")
        self.relic_search_entry.bind("<KeyRelease>", lambda e: self.filter_relics())
        ttk.Button(toolbar, text="Refresh", command=self.refresh).pack(
            side="right", padx=5
        )

        main_frame = ttk.Frame(self.relics_frame)
        main_frame.pack(fill="both", expand=True, padx=5, pady=5)

        columns = ("Name", "ID", "Runs", "Wins", "Win%", "Description")
        self.relic_tree = ttk.Treeview(
            main_frame, columns=columns, show="headings", selectmode="browse"
        )

        self.relic_tree.heading(
            "Name", text="Relic Name", command=lambda: self.sort_relics("name")
        )
        self.relic_tree.heading(
            "ID", text="Relic ID", command=lambda: self.sort_relics("id")
        )
        self.relic_tree.heading(
            "Runs", text="Runs", command=lambda: self.sort_relics("total")
        )
        self.relic_tree.heading(
            "Wins", text="Wins", command=lambda: self.sort_relics("wins")
        )
        self.relic_tree.heading(
            "Win%", text="Win Rate %", command=lambda: self.sort_relics("win_rate")
        )
        self.relic_tree.heading(
            "Description",
            text="Description",
            command=lambda: self.sort_relics("description"),
        )

        self.relic_tree.column("Name", width=150)
        self.relic_tree.column("ID", width=150)
        self.relic_tree.column("Runs", width=60, anchor="e")
        self.relic_tree.column("Wins", width=60, anchor="e")
        self.relic_tree.column("Win%", width=60, anchor="e")
        self.relic_tree.column("Description", width=500)

        scrollbar_y = ttk.Scrollbar(
            main_frame, orient="vertical", command=self.relic_tree.yview
        )
        scrollbar_x = ttk.Scrollbar(
            main_frame, orient="horizontal", command=self.relic_tree.xview
        )
        self.relic_tree.configure(
            yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set
        )

        self.relic_tree.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")

        self.relic_tree.tag_configure("oddrow", background="#f0f0f0")
        self.relic_tree.tag_configure("evenrow", background="#ffffff")

        self.relic_desc_label = ttk.Label(
            self.relics_frame, text="", wraplength=700, justify="left"
        )
        self.relic_desc_label.pack(fill="x", padx=5, pady=2)
        self.relic_tree.bind("<<TreeviewSelect>>", self.on_relic_select)

    def create_events_tab(self):
        toolbar = ttk.Frame(self.events_frame)
        toolbar.pack(fill="x", padx=5, pady=5)

        ttk.Label(toolbar, text="Search:").pack(side="left", padx=5)
        self.event_search_var = tk.StringVar()
        self.event_search_entry = ttk.Entry(
            toolbar, textvariable=self.event_search_var, width=20
        )
        self.event_search_entry.pack(side="left")
        self.event_search_entry.bind("<KeyRelease>", lambda e: self.filter_events())
        ttk.Button(toolbar, text="Refresh", command=self.refresh).pack(
            side="right", padx=5
        )

        main_frame = ttk.Frame(self.events_frame)
        main_frame.pack(fill="both", expand=True, padx=5, pady=5)

        columns = ("Event", "Choice", "Runs", "Wins", "Win%", "Description")
        self.event_tree = ttk.Treeview(
            main_frame, columns=columns, show="headings", selectmode="browse"
        )

        self.event_tree.heading(
            "Event", text="Event ID", command=lambda: self.sort_events("event")
        )
        self.event_tree.heading(
            "Choice", text="Choice", command=lambda: self.sort_events("choice")
        )
        self.event_tree.heading(
            "Runs", text="Times", command=lambda: self.sort_events("total")
        )
        self.event_tree.heading(
            "Wins", text="Wins", command=lambda: self.sort_events("wins")
        )
        self.event_tree.heading(
            "Win%", text="Win Rate %", command=lambda: self.sort_events("win_rate")
        )
        self.event_tree.heading(
            "Description",
            text="Description",
            command=lambda: self.sort_events("description"),
        )

        self.event_tree.column("Event", width=150)
        self.event_tree.column("Choice", width=150)
        self.event_tree.column("Runs", width=60, anchor="e")
        self.event_tree.column("Wins", width=60, anchor="e")
        self.event_tree.column("Win%", width=60, anchor="e")
        self.event_tree.column("Description", width=500)

        scrollbar_y = ttk.Scrollbar(
            main_frame, orient="vertical", command=self.event_tree.yview
        )
        scrollbar_x = ttk.Scrollbar(
            main_frame, orient="horizontal", command=self.event_tree.xview
        )
        self.event_tree.configure(
            yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set
        )

        self.event_tree.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")

        self.event_tree.tag_configure("oddrow", background="#f0f0f0")
        self.event_tree.tag_configure("evenrow", background="#ffffff")

        self.event_desc_label = ttk.Label(
            self.events_frame, text="", wraplength=700, justify="left"
        )
        self.event_desc_label.pack(fill="x", padx=5, pady=2)
        self.event_tree.bind("<<TreeviewSelect>>", self.on_event_select)

    def create_ancient_tab(self):
        toolbar = ttk.Frame(self.ancient_frame)
        toolbar.pack(fill="x", padx=5, pady=5)

        ttk.Label(toolbar, text="Search:").pack(side="left", padx=5)
        self.ancient_search_var = tk.StringVar()
        self.ancient_search_entry = ttk.Entry(
            toolbar, textvariable=self.ancient_search_var, width=20
        )
        self.ancient_search_entry.pack(side="left")
        self.ancient_search_entry.bind("<KeyRelease>", lambda e: self.filter_ancient())
        ttk.Button(toolbar, text="Refresh", command=self.refresh).pack(
            side="right", padx=5
        )

        main_frame = ttk.Frame(self.ancient_frame)
        main_frame.pack(fill="both", expand=True, padx=5, pady=5)

        columns = (
            "Event",
            "Relic",
            "Offered",
            "Picked",
            "Pick%",
            "Wins",
            "Win%",
            "Description",
        )
        self.ancient_tree = ttk.Treeview(
            main_frame, columns=columns, show="headings", selectmode="browse"
        )

        self.ancient_tree.heading(
            "Event", text="Event", command=lambda: self.sort_ancient("event")
        )
        self.ancient_tree.heading(
            "Relic", text="Relic", command=lambda: self.sort_ancient("relic")
        )
        self.ancient_tree.heading(
            "Offered", text="Offered", command=lambda: self.sort_ancient("offered")
        )
        self.ancient_tree.heading(
            "Picked", text="Picked", command=lambda: self.sort_ancient("picked")
        )
        self.ancient_tree.heading(
            "Pick%", text="Pick%", command=lambda: self.sort_ancient("pick_rate")
        )
        self.ancient_tree.heading(
            "Wins", text="Wins", command=lambda: self.sort_ancient("wins")
        )
        self.ancient_tree.heading(
            "Win%", text="Win%", command=lambda: self.sort_ancient("win_rate")
        )
        self.ancient_tree.heading(
            "Description",
            text="Description",
            command=lambda: self.sort_ancient("description"),
        )

        self.ancient_tree.column("Event", width=120)
        self.ancient_tree.column("Relic", width=150)
        self.ancient_tree.column("Offered", width=60, anchor="e")
        self.ancient_tree.column("Picked", width=60, anchor="e")
        self.ancient_tree.column("Pick%", width=60, anchor="e")
        self.ancient_tree.column("Wins", width=50, anchor="e")
        self.ancient_tree.column("Win%", width=60, anchor="e")
        self.ancient_tree.column("Description", width=400)

        scrollbar_y = ttk.Scrollbar(
            main_frame, orient="vertical", command=self.ancient_tree.yview
        )
        scrollbar_x = ttk.Scrollbar(
            main_frame, orient="horizontal", command=self.ancient_tree.xview
        )
        self.ancient_tree.configure(
            yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set
        )

        self.ancient_tree.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")

        self.ancient_tree.tag_configure("oddrow", background="#f0f0f0")
        self.ancient_tree.tag_configure("evenrow", background="#ffffff")

        self.ancient_desc_label = ttk.Label(
            self.ancient_frame, text="", wraplength=700, justify="left"
        )
        self.ancient_desc_label.pack(fill="x", padx=5, pady=2)
        self.ancient_tree.bind("<<TreeviewSelect>>", self.on_ancient_select)

    def load_ancient(self):
        self.all_ancient_data = self.ancient_data
        self.filter_ancient()

    def filter_ancient(self):
        search = self.ancient_search_var.get().lower()
        filtered = self.all_ancient_data
        if search:
            filtered = [
                e
                for e in self.all_ancient_data
                if search in e["event"].lower() or search in e["relic"].lower()
            ]

        reverse = getattr(self, "_ancient_sort_reverse", False)
        key = getattr(self, "_ancient_sort_key", "offered")

        def get_val(item):
            val = item.get(key)
            if isinstance(val, str):
                return val.lower()
            return val if val else 0

        filtered = sorted(filtered, key=get_val, reverse=reverse)

        for row in self.ancient_tree.get_children():
            self.ancient_tree.delete(row)

        for i, item in enumerate(filtered):
            tags = ("oddrow",) if i % 2 == 0 else ("evenrow",)
            self.ancient_tree.insert(
                "",
                "end",
                values=(
                    item["event"],
                    item["relic"],
                    item["offered"],
                    item["picked"],
                    item["pick_rate"],
                    item["wins"],
                    item["win_rate"],
                    item["description"],
                ),
                tags=tags,
            )

    def sort_ancient(self, key):
        reverse = getattr(self, "_ancient_sort_reverse", False)
        self._ancient_sort_reverse = not reverse
        self._ancient_sort_key = key
        self.filter_ancient()

    def on_ancient_select(self, event):
        selection = self.ancient_tree.selection()
        if selection:
            idx = self.ancient_tree.index(selection[0])
            if idx < len(self.ancient_data):
                item = self.ancient_data[idx]
                self.ancient_desc_label.config(
                    text=f"{item['relic']}: {item['description']}"
                )

    def create_encounters_tab(self):
        toolbar = ttk.Frame(self.encounters_frame)
        toolbar.pack(fill="x", padx=5, pady=5)

        ttk.Label(toolbar, text="Search:").pack(side="left", padx=5)
        self.encounters_search_var = tk.StringVar()
        self.encounters_search_entry = ttk.Entry(
            toolbar, textvariable=self.encounters_search_var, width=20
        )
        self.encounters_search_entry.pack(side="left")
        self.encounters_search_entry.bind(
            "<KeyRelease>", lambda e: self.filter_encounters()
        )
        ttk.Button(toolbar, text="Refresh", command=self.refresh).pack(
            side="right", padx=5
        )

        main_frame = ttk.Frame(self.encounters_frame)
        main_frame.pack(fill="both", expand=True, padx=5, pady=5)

        columns = (
            "Encounter",
            "Act#",
            "Enemy",
            "Times",
            "Total Dmg",
            "Avg Dmg",
            "Wins",
            "Win%",
        )
        self.encounters_tree = ttk.Treeview(
            main_frame, columns=columns, show="headings", selectmode="browse"
        )

        self.encounters_tree.heading("Encounter", text="Encounter")
        self.encounters_tree.heading("Act#", text="Act#")
        self.encounters_tree.heading("Enemy", text="Enemy")
        self.encounters_tree.heading("Times", text="Times")
        self.encounters_tree.heading("Total Dmg", text="Total Damage")
        self.encounters_tree.heading("Avg Dmg", text="Avg Damage")
        self.encounters_tree.heading("Wins", text="Wins")
        self.encounters_tree.heading("Win%", text="Win%")

        self.encounters_tree.column("Encounter", width=150)
        self.encounters_tree.column("Act#", width=50, anchor="e")
        self.encounters_tree.column("Enemy", width=280)
        self.encounters_tree.column("Times", width=60, anchor="e")
        self.encounters_tree.column("Total Dmg", width=80, anchor="e")
        self.encounters_tree.column("Avg Dmg", width=80, anchor="e")
        self.encounters_tree.column("Wins", width=60, anchor="e")
        self.encounters_tree.column("Win%", width=60, anchor="e")

        scrollbar_y = ttk.Scrollbar(
            main_frame, orient="vertical", command=self.encounters_tree.yview
        )
        self.encounters_tree.configure(yscrollcommand=scrollbar_y.set)
        self.encounters_tree.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")

        self.encounters_tree.tag_configure("oddrow", background="#f0f0f0")
        self.encounters_tree.tag_configure("evenrow", background="#ffffff")

        self.encounters_sort_col = "Times"
        self.encounters_sort_rev = True
        for col in columns:
            self.encounters_tree.heading(
                col, text=col, command=lambda c=col: self.sort_encounters(c)
            )

    def load_encounters(self):
        self.all_encounters_data = self.encounters_data
        self.filter_encounters()

    def filter_encounters(self):
        search = self.encounters_search_var.get().lower()
        filtered = self.all_encounters_data
        if search:
            filtered = [
                e
                for e in self.all_encounters_data
                if search in e["encounter"].lower()
                or search in e.get("enemy", "").lower()
            ]

        col = self.encounters_sort_col
        reverse = self.encounters_sort_rev
        col_map = {
            "Encounter": "encounter",
            "Act#": "act",
            "Enemy": "enemy",
            "Times": "times",
            "Total Dmg": "total_damage",
            "Avg Dmg": "avg_damage",
            "Wins": "wins",
            "Win%": "win_rate",
        }
        key_col = col_map.get(col, "times")
        filtered = sorted(
            filtered, key=lambda x: self._sort_key(x, key_col), reverse=reverse
        )

        for row in self.encounters_tree.get_children():
            self.encounters_tree.delete(row)

        for i, item in enumerate(filtered):
            tags = ("oddrow",) if i % 2 == 0 else ("evenrow",)
            self.encounters_tree.insert(
                "",
                "end",
                values=(
                    item["encounter"],
                    item.get("act", "-"),
                    item.get("enemy", "-"),
                    item["times"],
                    item["total_damage"],
                    item["avg_damage"],
                    item["wins"],
                    item["win_rate"],
                ),
                tags=tags,
            )

    def _sort_key(self, item, key):
        val = item.get(key)
        if key == "enemy":
            return val.lower() if val else ""
        if key == "act":
            return val if val else 0
        if isinstance(val, str):
            return val.lower()
        return val if val else 0

    def sort_encounters(self, col):
        if self.encounters_sort_col == col:
            self.encounters_sort_rev = not self.encounters_sort_rev
        else:
            self.encounters_sort_col = col
            self.encounters_sort_rev = True
        self.filter_encounters()

    def find_and_load_data(self):
        self.history_dir = find_sts2_history_dir()
        if self.history_dir:
            self.path_label.config(text=self.history_dir, foreground="black")
            self.generate_data()
        else:
            self.path_label.config(
                text="Not found - click Browse to select", foreground="red"
            )

    def browse_folder(self):
        folder = filedialog.askdirectory(title="Select STS2 save folder")
        if folder:
            history_path = os.path.join(folder, "saves", "history")
            if os.path.exists(history_path) and glob.glob(
                os.path.join(history_path, "*.run")
            ):
                self.history_dir = history_path
                self.path_label.config(text=history_path, foreground="black")
                self.generate_data()
            else:
                self.history_dir = folder
                self.path_label.config(text=folder, foreground="black")
                self.generate_data()

    def generate_data(self):
        if not self.history_dir:
            messagebox.showwarning("Warning", "Please select a STS2 save folder first.")
            return

        self.status_label.config(text="Loading runs...")
        self.update()

        runs = load_runs(self.history_dir)
        if not runs:
            messagebox.showwarning(
                "No Data", "No run files found in the selected folder."
            )
            self.status_label.config(text="No runs found")
            return

        self.status_label.config(text=f"Processing {len(runs)} runs...")
        self.update()

        pick_by_class, win_by_class = get_card_data(runs)
        relic_stats = get_relic_data(runs)
        event_stats = get_event_data(runs)
        ancient_stats = get_ancient_relic_data(runs)
        encounter_stats = get_encounter_data(runs)
        total = create_excel(
            pick_by_class,
            win_by_class,
            relic_stats,
            event_stats,
            ancient_stats,
            encounter_stats,
            EXCEL_FILE,
        )

        self.status_label.config(
            text=f"Generated {total} cards, {len(relic_stats)} relics, {sum(len(e) for e in event_stats.values())} events, {len(ancient_stats)} ancient relics, {len(encounter_stats)} encounters from {len(runs)} runs"
        )
        self.load_excel_data()

    def load_excel_data(self):
        if not os.path.exists(EXCEL_FILE):
            return

        self.card_data = {}
        self.relic_data = []
        self.event_data = []
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            for class_name in wb.sheetnames:
                ws = wb[class_name]
                if class_name == "Relics":
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            by_class = {}
                            for i, cls in enumerate(CLASS_COLORS.keys()):
                                total = row[7 + i * 3] or 0
                                wins = row[7 + i * 3 + 1] or 0
                                by_class[cls] = {"total": total, "wins": wins}
                            self.relic_data.append(
                                {
                                    "id": row[0],
                                    "name": row[1],
                                    "total": row[2],
                                    "wins": row[3],
                                    "win_rate": row[4],
                                    "description": row[6] if len(row) > 6 else "",
                                    "by_class": by_class,
                                }
                            )
                elif class_name == "Events":
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            by_class = {}
                            for i, cls in enumerate(CLASS_COLORS.keys()):
                                total = row[7 + i * 3] or 0
                                wins = row[7 + i * 3 + 1] or 0
                                by_class[cls] = {"total": total, "wins": wins}
                            self.event_data.append(
                                {
                                    "event": row[0],
                                    "choice": row[1],
                                    "total": row[2],
                                    "wins": row[3],
                                    "win_rate": row[4],
                                    "description": row[6] if len(row) > 6 else "",
                                    "by_class": by_class,
                                }
                            )
                elif class_name == "Ancient Relics":
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            self.ancient_data.append(
                                {
                                    "event": row[0],
                                    "relic": row[1],
                                    "offered": row[2],
                                    "picked": row[3],
                                    "pick_rate": row[4],
                                    "wins": row[5],
                                    "win_rate": row[6],
                                    "description": row[22] if len(row) > 22 else "",
                                }
                            )
                elif class_name == "Encounters":
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            self.encounters_data.append(
                                {
                                    "encounter": row[0],
                                    "act": row[1],
                                    "enemy": row[2],
                                    "times": row[3],
                                    "total_damage": row[4],
                                    "avg_damage": row[5],
                                    "wins": row[6],
                                    "win_rate": row[7],
                                }
                            )
                else:
                    self.card_data[class_name] = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            self.card_data[class_name].append(
                                {
                                    "card": row[0],
                                    "original_class": row[1],
                                    "offered": row[2],
                                    "picked": row[3],
                                    "pick_rate": row[4],
                                    "runs": row[5],
                                    "wins": row[6],
                                    "win_rate": row[7],
                                }
                            )
            self.load_class("IRONCLAD")
            self.load_relics()
            self.load_events()
            self.load_ancient()
            self.load_encounters()
            runs = load_runs(self.history_dir) if self.history_dir else []
            self.status_label.config(
                text=f"Loaded {sum(len(c) for c in self.card_data.values())} cards, {len(self.relic_data)} relics, {len(self.event_data)} events, {len(self.ancient_data)} ancient, {len(self.encounters_data)} encounters from {len(runs)} runs"
            )
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load data: {e}")

    def load_class(self, class_name):
        self.current_class = class_name
        self.all_data = self.card_data.get(class_name, [])
        self.filter_data()

    def filter_data(self):
        search = self.search_var.get().lower()
        filtered = self.all_data
        if search:
            filtered = [c for c in self.all_data if search in c["card"].lower()]

        for row in self.tree.get_children():
            self.tree.delete(row)

        for i, card in enumerate(filtered):
            tags = ("oddrow",) if i % 2 == 0 else ("evenrow",)
            self.tree.insert(
                "",
                "end",
                values=(
                    card["card"],
                    card["original_class"],
                    card["offered"],
                    card["picked"],
                    card["pick_rate"],
                    card["runs"],
                    card["wins"],
                    card["win_rate"],
                ),
                tags=tags,
            )

    def sort_by(self, key):
        reverse = getattr(self, "_sort_reverse", False)
        self._sort_reverse = not reverse

        def get_val(card):
            val = card.get(key)
            if isinstance(val, str):
                return val.lower()
            return val if val else 0

        self.all_data.sort(key=get_val, reverse=reverse)
        self.filter_data()

    def on_class_change(self):
        self.load_class(self.class_var.get())

    def load_relics(self):
        self.filter_relics()

    def load_events(self):
        self.all_event_data = self.event_data
        self.filter_events()

    def filter_events(self):
        search = self.event_search_var.get().lower()
        filtered = self.all_event_data
        if search:
            filtered = [
                e
                for e in self.all_event_data
                if search in e["event"].lower() or search in e["choice"].lower()
            ]

        for row in self.event_tree.get_children():
            self.event_tree.delete(row)

        for i, event in enumerate(filtered):
            tags = ("oddrow",) if i % 2 == 0 else ("evenrow",)
            self.event_tree.insert(
                "",
                "end",
                values=(
                    event["event"],
                    event["choice"],
                    event["total"],
                    event["wins"],
                    event["win_rate"],
                    event["description"],
                ),
                tags=tags,
            )

    def sort_events(self, key):
        reverse = getattr(self, "_event_sort_reverse", False)
        self._event_sort_reverse = not reverse

        def get_val(event):
            val = event.get(key)
            if isinstance(val, str):
                return val.lower()
            return val if val else 0

        self.all_event_data.sort(key=get_val, reverse=reverse)
        self.filter_events()

    def on_event_select(self, event):
        selection = self.event_tree.selection()
        if selection:
            item = self.event_tree.item(selection[0])
            idx = self.event_tree.index(selection[0])
            if idx < len(self.event_data):
                ev = self.event_data[idx]
                class_stats = []
                for cls, stats in ev.get("by_class", {}).items():
                    total = stats.get("total", 0)
                    if isinstance(total, str):
                        total = 0
                    if total > 0:
                        wins = stats.get("wins", 0)
                        win_rate = (wins / total * 100) if total > 0 else 0
                        class_stats.append(f"{cls}: {total} ({wins}W, {win_rate:.0f}%)")
                class_text = " | ".join(class_stats) if class_stats else "No class data"
                self.event_desc_label.config(
                    text=f"{class_text}\n{ev.get('description', '')}"
                )

    def filter_relics(self):
        search = self.relic_search_var.get().lower()
        filtered = self.relic_data
        if search:
            filtered = [
                r
                for r in self.relic_data
                if search in r["name"].lower() or search in r["id"].lower()
            ]

        for row in self.relic_tree.get_children():
            self.relic_tree.delete(row)

        for i, relic in enumerate(filtered):
            tags = ("oddrow",) if i % 2 == 0 else ("evenrow",)
            self.relic_tree.insert(
                "",
                "end",
                values=(
                    relic["name"],
                    relic["id"],
                    relic["total"],
                    relic["wins"],
                    relic["win_rate"],
                    relic["description"],
                ),
                tags=tags,
            )

    def sort_relics(self, key):
        reverse = getattr(self, "_relic_sort_reverse", False)
        self._relic_sort_reverse = not reverse

        def get_val(relic):
            val = relic.get(key)
            if isinstance(val, str):
                return val.lower()
            return val if val else 0

        self.relic_data.sort(key=get_val, reverse=reverse)
        self.filter_relics()

    def on_relic_select(self, event):
        selection = self.relic_tree.selection()
        if selection:
            item = self.relic_tree.item(selection[0])
            idx = self.relic_tree.index(selection[0])
            if idx < len(self.relic_data):
                relic = self.relic_data[idx]
                class_stats = []
                for cls, stats in relic.get("by_class", {}).items():
                    total = stats.get("total", 0)
                    if isinstance(total, str):
                        total = 0
                    if total > 0:
                        wins = stats.get("wins", 0)
                        win_rate = (wins / total * 100) if total > 0 else 0
                        class_stats.append(f"{cls}: {total} ({wins}W, {win_rate:.0f}%)")
                class_text = " | ".join(class_stats) if class_stats else "No class data"
                self.relic_desc_label.config(
                    text=f"{class_text}\n{relic.get('description', '')}"
                )

    def refresh(self):
        self.find_and_load_data()

    def show_help(self):
        help_text = """How to find your STS2 save folder:

WINDOWS:
%LOCALAPPDATA%\\SlayTheSpire2\\steam\\<profile>\\saves\\history

MAC:
~/Library/Application Support/SlayTheSpire2/steam/<profile>/saves/history

LINUX:
~/.local/share/SlayTheSpire2/steam/<profile>/saves/history

NOTE: <profile> is usually a number like 76561198054269638

If auto-detection fails, click "Browse" and select the folder containing your .run files."""
        messagebox.showinfo("Help - Finding STS2 Save Files", help_text)


if __name__ == "__main__":
    app = STSCardViewer()
    app.mainloop()
