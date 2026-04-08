#!/usr/bin/env python3
import os
import sys
import glob
import json
import argparse
from collections import defaultdict, Counter
from data.card_classifications import CARD_CLASSIFICATIONS
from data.relic_info import RELIC_INFO
from data.event_info import EVENT_INFO
from data.event_choices import EVENT_CHOICES

EXCEL_FILE = "sts2_cards.xlsx"

CLASS_COLORS = {
    "IRONCLAD": "FF8B8B",
    "SILENT": "8BFF8B",
    "DEFECT": "8B8BFF",
    "NECROBINDER": "FF8BFF",
    "REGENT": "FFFF8B",
}


def find_sts2_history_dir():
    if sys.platform == "win32":
        paths = [
            os.path.join(os.environ.get("APPDATA", ""), "SlayTheSpire2", "steam"),
            os.path.join(os.environ.get("LOCALAPPDATA", ""), "SlayTheSpire2", "steam"),
        ]
    elif sys.platform == "darwin":
        paths = [
            os.path.expanduser("~/Library/Application Support/SlayTheSpire2/steam/")
        ]
    else:
        paths = [os.path.expanduser("~/.local/share/SlayTheSpire2/steam/")]

    for base_path in paths:
        if not os.path.exists(base_path):
            continue
        try:
            for entry in os.listdir(base_path):
                profile_base = os.path.join(base_path, entry)
                for subpath in ["profile1/saves/history", "saves/history"]:
                    profile_path = os.path.join(profile_base, subpath)
                    if os.path.isdir(profile_path) and glob.glob(
                        os.path.join(profile_path, "*.run")
                    ):
                        return profile_path
        except PermissionError:
            continue
    return None


def load_runs(history_dir):
    runs = []
    for f in sorted(glob.glob(os.path.join(history_dir, "*.run"))):
        try:
            with open(f, "r") as file:
                runs.append(json.load(file))
        except:
            continue
    return runs


def get_card_data(runs):
    pick_by_class = defaultdict(
        lambda: defaultdict(lambda: {"offered": 0, "picked": 0})
    )
    win_by_class = defaultdict(lambda: defaultdict(lambda: {"runs": 0, "wins": 0}))

    for run in runs:
        is_win = run.get("win", False)
        players = run.get("players", [])
        if not players:
            continue
        char = players[0].get("character", "Unknown").replace("CHARACTER.", "")
        if char not in CLASS_COLORS:
            continue

        cards_in_run = set()
        for act in run.get("map_point_history", []):
            for point in act:
                for stat in point.get("player_stats", []):
                    for choice in stat.get("card_choices", []):
                        card_id = (
                            choice.get("card", {})
                            .get("id", "Unknown")
                            .replace("CARD.", "")
                        )
                        was_picked = choice.get("was_picked", False)
                        pick_by_class[char][card_id]["offered"] += 1
                        if was_picked:
                            pick_by_class[char][card_id]["picked"] += 1
                            cards_in_run.add(card_id)

        for card_id in cards_in_run:
            win_by_class[char][card_id]["runs"] += 1
            if is_win:
                win_by_class[char][card_id]["wins"] += 1

    return pick_by_class, win_by_class


def get_relic_data(runs):
    relic_stats = defaultdict(
        lambda: {
            "total": 0,
            "wins": 0,
            "by_class": defaultdict(lambda: {"total": 0, "wins": 0}),
        }
    )
    for run in runs:
        is_win = run.get("win", False)
        players = run.get("players", [])
        if not players:
            continue
        char = players[0].get("character", "Unknown").replace("CHARACTER.", "")
        if char not in CLASS_COLORS:
            continue
        for relic in players[0].get("relics", []):
            rid = relic.get("id", "Unknown").replace("RELIC.", "")
            relic_stats[rid]["total"] += 1
            relic_stats[rid]["by_class"][char]["total"] += 1
            if is_win:
                relic_stats[rid]["wins"] += 1
                relic_stats[rid]["by_class"][char]["wins"] += 1
    return relic_stats


def get_event_data(runs):
    event_stats = defaultdict(
        lambda: defaultdict(
            lambda: {
                "total": 0,
                "wins": 0,
                "encounters": 0,
                "is_ancient": False,
                "by_class": defaultdict(
                    lambda: {"total": 0, "wins": 0, "encounters": 0}
                ),
            }
        )
    )
    for run in runs:
        is_win = run.get("win", False)
        players = run.get("players", [])
        if not players:
            continue
        char = players[0].get("character", "Unknown").replace("CHARACTER.", "")
        if char not in CLASS_COLORS:
            continue
        for act in run.get("map_point_history", []):
            for point in act:
                rooms = point.get("rooms", [])
                if not rooms:
                    continue
                room = rooms[0]
                if room.get("room_type") != "event":
                    continue
                event_id = room.get("model_id", "").replace("EVENT.", "")
                is_ancient = point.get("map_point_type") == "ancient"
                event_stats[event_id]["_meta"]["encounters"] += 1
                event_stats[event_id]["_meta"]["is_ancient"] = is_ancient
                event_stats[event_id]["_meta"]["by_class"][char]["encounters"] += 1
                if is_win:
                    event_stats[event_id]["_meta"]["wins"] += 1
                    event_stats[event_id]["_meta"]["by_class"][char]["wins"] += 1
                for stat in point.get("player_stats", []):
                    for choice in stat.get("event_choices", []):
                        title_key = choice.get("title", {}).get("key", "")
                        if "options." in title_key:
                            choice_id = title_key.split("options.")[-1].split(".")[0]
                        else:
                            choice_id = "DEFAULT"
                        event_stats[event_id][choice_id]["total"] += 1
                        event_stats[event_id][choice_id]["by_class"][char]["total"] += 1
                        if is_win:
                            event_stats[event_id][choice_id]["wins"] += 1
                            event_stats[event_id][choice_id]["by_class"][char][
                                "wins"
                            ] += 1
    return event_stats


def get_ancient_relic_data(runs):
    ancient_stats = defaultdict(
        lambda: defaultdict(
            lambda: {
                "offered": 0,
                "picked": 0,
                "wins": 0,
                "by_class": defaultdict(lambda: {"offered": 0, "picked": 0, "wins": 0}),
            }
        )
    )
    invalid_keys = {"IRONCLAD", "SILENT", "DEFECT", "NECROBINDER", "REGENT", ""}
    for run in runs:
        is_win = run.get("win", False)
        players = run.get("players", [])
        if not players:
            continue
        char = players[0].get("character", "Unknown").replace("CHARACTER.", "")
        if char not in CLASS_COLORS:
            continue
        for act in run.get("map_point_history", []):
            for point in act:
                if point.get("map_point_type") != "ancient":
                    continue
                rooms = point.get("rooms", [])
                if not rooms:
                    continue
                event_id = rooms[0].get("model_id", "").replace("EVENT.", "")
                for stat in point.get("player_stats", []):
                    for ancient_choice in stat.get("ancient_choice", []):
                        rid = ancient_choice.get("TextKey", "").replace("RELIC.", "")
                        if rid in invalid_keys:
                            continue
                        was_chosen = ancient_choice.get("was_chosen", False)
                        ancient_stats[event_id][rid]["offered"] += 1
                        ancient_stats[event_id][rid]["by_class"][char]["offered"] += 1
                        if was_chosen:
                            ancient_stats[event_id][rid]["picked"] += 1
                            ancient_stats[event_id][rid]["by_class"][char][
                                "picked"
                            ] += 1
                        if was_chosen and is_win:
                            ancient_stats[event_id][rid]["wins"] += 1
                            ancient_stats[event_id][rid]["by_class"][char]["wins"] += 1
    return ancient_stats


def get_encounter_data(runs):
    """Track damage taken from each encounter"""
    encounter_stats = defaultdict(
        lambda: {
            "count": 0,
            "total_damage": 0,
            "wins": 0,
            "acts": defaultdict(int),
            "enemies": Counter(),
        }
    )
    for run in runs:
        is_win = run.get("win", False)
        for act_idx, act in enumerate(run.get("map_point_history", [])):
            for point in act:
                rooms = point.get("rooms", [])
                if not rooms:
                    continue
                room = rooms[0]
                room_id = room.get("model_id", "")
                if "ENCOUNTER" not in room_id:
                    continue
                encounter_id = room_id.replace("ENCOUNTER.", "")
                monster_ids = room.get("monster_ids", [])
                for stat in point.get("player_stats", []):
                    damage = stat.get("damage_taken", 0)
                    encounter_stats[encounter_id]["count"] += 1
                    encounter_stats[encounter_id]["total_damage"] += damage
                    encounter_stats[encounter_id]["acts"][act_idx + 1] += 1
                    for m in monster_ids:
                        m_name = m.replace("MONSTER.", "")
                        encounter_stats[encounter_id]["enemies"][m_name] += 1
                    if is_win:
                        encounter_stats[encounter_id]["wins"] += 1
    return encounter_stats


def create_excel(
    pick_by_class,
    win_by_class,
    relic_stats,
    event_stats,
    ancient_stats,
    encounter_stats,
    output_path,
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for class_name in CLASS_COLORS.keys():
        ws = wb.create_sheet(title=class_name)
        header_fill = PatternFill(
            start_color="FF" + CLASS_COLORS[class_name],
            end_color="FF" + CLASS_COLORS[class_name],
            fill_type="solid",
        )
        headers = [
            "Card",
            "Original Class",
            "Times Offered",
            "Times Picked",
            "Pick Rate %",
            "Runs with Card",
            "Wins with Card",
            "Win Rate %",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        row = 2
        class_cards = list(pick_by_class[class_name].items())
        class_cards.sort(key=lambda x: x[1]["picked"], reverse=True)

        for card_id, pick_stats in class_cards:
            win_stats = win_by_class[class_name].get(card_id, {"runs": 0, "wins": 0})
            original_class = CARD_CLASSIFICATIONS.get(card_id, "COLORLESS")
            pick_rate = (
                (pick_stats["picked"] / pick_stats["offered"] * 100)
                if pick_stats["offered"] > 0
                else 0
            )
            win_rate = (
                (win_stats["wins"] / win_stats["runs"] * 100)
                if win_stats["runs"] > 0
                else 0
            )

            ws.cell(row=row, column=1, value=card_id)
            ws.cell(row=row, column=2, value=original_class)
            ws.cell(row=row, column=3, value=pick_stats["offered"])
            ws.cell(row=row, column=4, value=pick_stats["picked"])
            ws.cell(row=row, column=5, value=round(pick_rate, 1))
            ws.cell(row=row, column=6, value=win_stats["runs"])
            ws.cell(row=row, column=7, value=win_stats["wins"])
            ws.cell(row=row, column=8, value=round(win_rate, 1))
            row += 1

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        for col in range(3, 9):
            ws.column_dimensions[get_column_letter(col)].width = 14

    # Relics sheet
    ws = wb.create_sheet(title="Relics")
    header_fill = PatternFill(
        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
    )
    headers = ["ID", "Name", "Runs", "Wins", "Win%", "", "Description"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for i, class_name in enumerate(CLASS_COLORS.keys()):
        ws.cell(row=1, column=8 + i * 3, value=class_name)
        ws.cell(row=1, column=8 + i * 3 + 1, value="W")
        ws.cell(row=1, column=8 + i * 3 + 2, value="%")

    row = 2
    relics_list = [(rid, stats) for rid, stats in relic_stats.items()]
    relics_list.sort(key=lambda x: x[1]["total"], reverse=True)

    for relic_id, stats in relics_list:
        win_rate = (stats["wins"] / stats["total"] * 100) if stats["total"] > 0 else 0
        relic_info = RELIC_INFO.get(
            relic_id, {"name": relic_id, "description": "Unknown"}
        )

        ws.cell(row=row, column=1, value=relic_id)
        ws.cell(row=row, column=2, value=relic_info.get("name", relic_id))
        ws.cell(row=row, column=3, value=stats["total"])
        ws.cell(row=row, column=4, value=stats["wins"])
        ws.cell(row=row, column=5, value=round(win_rate, 1))
        ws.cell(row=row, column=6, value="")
        ws.cell(row=row, column=7, value=relic_info.get("description", ""))

        for i, class_name in enumerate(CLASS_COLORS.keys()):
            class_stats = stats["by_class"].get(class_name, {"total": 0, "wins": 0})
            class_win_rate = (
                (class_stats["wins"] / class_stats["total"] * 100)
                if class_stats["total"] > 0
                else 0
            )
            ws.cell(row=row, column=8 + i * 3, value=class_stats["total"])
            ws.cell(row=row, column=8 + i * 3 + 1, value=class_stats["wins"])
            ws.cell(
                row=row,
                column=8 + i * 3 + 2,
                value=round(class_win_rate, 1) if class_stats["total"] > 0 else "-",
            )
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 2
    ws.column_dimensions["G"].width = 45
    for i in range(len(CLASS_COLORS)):
        ws.column_dimensions[get_column_letter(8 + i * 3)].width = 8
        ws.column_dimensions[get_column_letter(8 + i * 3 + 1)].width = 5
        ws.column_dimensions[get_column_letter(8 + i * 3 + 2)].width = 5

    # Events sheet
    ws = wb.create_sheet(title="Events")
    header_fill = PatternFill(
        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
    )
    headers = ["Event", "Choice", "Encounters", "Picked", "Pick%", "Win%", "Wins"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for i, class_name in enumerate(CLASS_COLORS.keys()):
        ws.cell(row=1, column=8 + i * 3, value=class_name)
        ws.cell(row=1, column=8 + i * 3 + 1, value="W")
        ws.cell(row=1, column=8 + i * 3 + 2, value="%")

    ws.cell(row=1, column=23, value="Description")
    ws.cell(row=1, column=23).fill = header_fill
    ws.cell(row=1, column=23).font = Font(bold=True)

    row = 2
    regular_events = []

    for event_id, choices in event_stats.items():
        if event_id == "_meta":
            continue
        meta = event_stats[event_id].get(
            "_meta", {"encounters": 0, "wins": 0, "is_ancient": False}
        )
        is_ancient = meta.get("is_ancient", False)
        if is_ancient:
            continue
        for choice, stats in choices.items():
            if choice != "_meta":
                regular_events.append((event_id, choice, stats, meta))

    regular_events.sort(key=lambda x: x[3]["encounters"], reverse=True)

    for event_id, choice, stats, meta in regular_events:
        win_rate = (stats["wins"] / stats["total"] * 100) if stats["total"] > 0 else 0
        pick_rate = (
            (stats["total"] / meta["encounters"] * 100) if meta["encounters"] > 0 else 0
        )
        event_info = EVENT_INFO.get(
            event_id, {"name": event_id, "description": "Unknown"}
        )

        choice_choices = EVENT_CHOICES.get(event_id, {})
        choice_desc = choice_choices.get(choice, choice)
        if choice == "DEFAULT":
            choice_desc = event_info.get("description", "Default")[:50]

        ws.cell(row=row, column=1, value=event_id)
        ws.cell(row=row, column=2, value=choice_desc)
        ws.cell(row=row, column=3, value=meta["encounters"])
        ws.cell(row=row, column=4, value=stats["total"])
        ws.cell(row=row, column=5, value=round(pick_rate, 1))
        ws.cell(row=row, column=6, value=round(win_rate, 1))
        ws.cell(row=row, column=7, value=stats["wins"])

        for i, class_name in enumerate(CLASS_COLORS.keys()):
            class_stats = stats["by_class"].get(class_name, {"total": 0, "wins": 0})
            class_meta = meta["by_class"].get(class_name, {"encounters": 0, "wins": 0})
            class_pick_rate = (
                (class_stats["total"] / class_meta["encounters"] * 100)
                if class_meta["encounters"] > 0
                else 0
            )
            ws.cell(
                row=row,
                column=8 + i * 3,
                value=round(class_pick_rate, 0) if class_stats["total"] > 0 else "-",
            )
            ws.cell(row=row, column=8 + i * 3 + 1, value=class_stats["wins"])
            ws.cell(row=row, column=8 + i * 3 + 2, value="-")

        desc_cell = ws.cell(row=row, column=23, value=event_info.get("description", ""))
        desc_cell.fill = PatternFill(
            start_color="FFFACD", end_color="FFFACD", fill_type="solid"
        )
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 6
    for i in range(len(CLASS_COLORS)):
        ws.column_dimensions[get_column_letter(8 + i * 3)].width = 7
        ws.column_dimensions[get_column_letter(8 + i * 3 + 1)].width = 5
        ws.column_dimensions[get_column_letter(8 + i * 3 + 2)].width = 5
    ws.column_dimensions["W"].width = 5
    ws.column_dimensions["X"].width = 5
    ws.column_dimensions["Y"].width = 50

    # Ancient Relics sheet
    ws = wb.create_sheet(title="Ancient Relics")
    header_fill = PatternFill(
        start_color="FFD700", end_color="FFD700", fill_type="solid"
    )
    headers = ["Event", "Relic", "Offered", "Picked", "Pick%", "Wins", "Win%"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for i, class_name in enumerate(CLASS_COLORS.keys()):
        ws.cell(row=1, column=8 + i * 3, value=class_name)
        ws.cell(row=1, column=8 + i * 3 + 1, value="Off")
        ws.cell(row=1, column=8 + i * 3 + 2, value="Pick%")

    ws.cell(row=1, column=23, value="Relic Description")
    ws.cell(row=1, column=23).fill = header_fill
    ws.cell(row=1, column=23).font = Font(bold=True)

    row = 2
    ancient_list = []
    for event_id, relics in ancient_stats.items():
        for relic_id, stats in relics.items():
            ancient_list.append((event_id, relic_id, stats))
    ancient_list.sort(key=lambda x: x[2]["offered"], reverse=True)

    for event_id, relic_id, stats in ancient_list:
        pick_rate = (
            (stats["picked"] / stats["offered"] * 100) if stats["offered"] > 0 else 0
        )
        win_rate = (stats["wins"] / stats["picked"] * 100) if stats["picked"] > 0 else 0
        relic_info = RELIC_INFO.get(
            relic_id, {"name": relic_id, "description": "Unknown"}
        )

        ws.cell(row=row, column=1, value=event_id)
        ws.cell(row=row, column=2, value=relic_info.get("name", relic_id))
        ws.cell(row=row, column=3, value=stats["offered"])
        ws.cell(row=row, column=4, value=stats["picked"])
        ws.cell(row=row, column=5, value=round(pick_rate, 1))
        ws.cell(row=row, column=6, value=stats["wins"])
        ws.cell(row=row, column=7, value=round(win_rate, 1))

        for i, class_name in enumerate(CLASS_COLORS.keys()):
            class_stats = stats["by_class"].get(class_name, {"offered": 0, "picked": 0})
            class_pick_rate = (
                (class_stats["picked"] / class_stats["offered"] * 100)
                if class_stats["offered"] > 0
                else 0
            )
            ws.cell(row=row, column=8 + i * 3, value=class_stats["offered"])
            ws.cell(row=row, column=8 + i * 3 + 1, value=class_stats["picked"])
            ws.cell(
                row=row,
                column=8 + i * 3 + 2,
                value=round(class_pick_rate, 0) if class_stats["offered"] > 0 else "-",
            )

        desc_cell = ws.cell(row=row, column=23, value=relic_info.get("description", ""))
        desc_cell.fill = PatternFill(
            start_color="FFFACD", end_color="FFFACD", fill_type="solid"
        )
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 8
    for i in range(len(CLASS_COLORS)):
        ws.column_dimensions[get_column_letter(8 + i * 3)].width = 6
        ws.column_dimensions[get_column_letter(8 + i * 3 + 1)].width = 6
        ws.column_dimensions[get_column_letter(8 + i * 3 + 2)].width = 7
    ws.column_dimensions["W"].width = 5
    ws.column_dimensions["X"].width = 5
    ws.column_dimensions["Y"].width = 50

    # Encounters sheet
    ws = wb.create_sheet(title="Encounters")
    header_fill = PatternFill(
        start_color="CC6666", end_color="CC6666", fill_type="solid"
    )
    headers = [
        "Encounter",
        "Act#",
        "Enemy Name",
        "Times",
        "Total Damage",
        "Avg Damage",
        "Wins",
        "Win%",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    row = 2
    encounters_list = [(eid, stats) for eid, stats in encounter_stats.items()]
    encounters_list.sort(key=lambda x: x[1]["count"], reverse=True)

    for encounter_id, stats in encounters_list:
        avg_damage = stats["total_damage"] / stats["count"] if stats["count"] > 0 else 0
        win_rate = (stats["wins"] / stats["count"] * 100) if stats["count"] > 0 else 0

        most_common_act = (
            max(stats["acts"].items(), key=lambda x: x[1])[0] if stats["acts"] else "-"
        )
        enemy_list = ", ".join(stats["enemies"].keys()) if stats["enemies"] else "-"

        ws.cell(row=row, column=1, value=encounter_id)
        ws.cell(row=row, column=2, value=most_common_act)
        ws.cell(row=row, column=3, value=enemy_list)
        ws.cell(row=row, column=4, value=stats["count"])
        ws.cell(row=row, column=5, value=stats["total_damage"])
        ws.cell(row=row, column=6, value=round(avg_damage, 1))
        ws.cell(row=row, column=7, value=stats["wins"])
        ws.cell(row=row, column=8, value=round(win_rate, 1))
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 10

    wb.save(output_path)
    return sum(len(cards) for cards in pick_by_class.values())


def main():
    parser = argparse.ArgumentParser(description="STS2 Statistics Generator (No GUI)")
    parser.add_argument(
        "-o", "--output", default=EXCEL_FILE, help="Output Excel file path"
    )
    parser.add_argument(
        "-d", "--dir", help="STS2 history directory (auto-detected if not provided)"
    )
    args = parser.parse_args()

    history_dir = args.dir if args.dir else find_sts2_history_dir()
    if not history_dir:
        print("Error: Could not find STS2 save directory. Use -d to specify.")
        sys.exit(1)

    print(f"Loading runs from: {history_dir}")
    runs = load_runs(history_dir)
    print(f"Loaded {len(runs)} runs")

    if not runs:
        print("No runs found!")
        sys.exit(1)

    print("Processing data...")
    pick_by_class, win_by_class = get_card_data(runs)
    relic_stats = get_relic_data(runs)
    event_stats = get_event_data(runs)
    ancient_stats = get_ancient_relic_data(runs)
    encounter_stats = get_encounter_data(runs)

    print(f"Generating Excel: {args.output}")
    total = create_excel(
        pick_by_class,
        win_by_class,
        relic_stats,
        event_stats,
        ancient_stats,
        encounter_stats,
        args.output,
    )

    print(
        f"Done! Generated {total} cards, {len(relic_stats)} relics, {sum(len(e) for e in event_stats.values())} events, {len(ancient_stats)} ancient relics, {len(encounter_stats)} encounters"
    )


if __name__ == "__main__":
    main()
